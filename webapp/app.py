from flask import Flask, render_template, request, send_file, jsonify, abort
from defusedxml import ElementTree as ET
from collections import Counter, defaultdict
from datetime import datetime, timedelta
import os
import re
import base64
import gzip
import hashlib
import random
import zlib
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from pathlib import Path
import zipfile
import csv

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 200 * 1024 * 1024  # 200 MB

APP_VERSION = "v0.161"
app.jinja_env.globals["APP_VERSION"] = APP_VERSION


def _parse_bool(value, default=None):
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    normalized = str(value).strip().lower()
    if normalized in ("1", "true", "yes", "y", "on", "enabled", "enable"):
        return True
    if normalized in ("0", "false", "no", "n", "off", "disabled", "disable"):
        return False
    return default


SCRIPT_ACCESS_CODE_DEFAULT_ENABLED = _parse_bool(os.getenv("SCRIPT_ACCESS_CODE_ENABLED"), True)
SCRIPT_ACCESS_CODE_ENABLED = SCRIPT_ACCESS_CODE_DEFAULT_ENABLED
SCRIPT_ACCESS_CODE_API_DEFAULT_ENABLED = _parse_bool(os.getenv("SCRIPT_ACCESS_CODE_API_ENABLED"), True)
SCRIPT_ACCESS_CODE_API_ENABLED = SCRIPT_ACCESS_CODE_API_DEFAULT_ENABLED
DEPLOY_DOWNLOAD_ENABLED = _parse_bool(os.getenv("DEPLOY_DOWNLOAD_ENABLED"), True)


def _is_script_access_code_enabled():
    return bool(SCRIPT_ACCESS_CODE_ENABLED)


def _is_script_access_code_api_enabled():
    return bool(SCRIPT_ACCESS_CODE_API_ENABLED)


def _is_deploy_download_enabled():
    return bool(DEPLOY_DOWNLOAD_ENABLED)


def _set_script_access_code_enabled(enabled):
    global SCRIPT_ACCESS_CODE_ENABLED
    SCRIPT_ACCESS_CODE_ENABLED = bool(enabled)
    return SCRIPT_ACCESS_CODE_ENABLED


@app.context_processor
def _inject_feature_flags():
    return {
        "SCRIPT_ACCESS_CODE_ENABLED": _is_script_access_code_enabled(),
        "DEPLOY_DOWNLOAD_ENABLED": _is_deploy_download_enabled(),
    }


# Zentraler Perzentilwert für die Performance-Auswertung.
# UI-Texte und Tabellenköpfe werden daraus dynamisch erzeugt.
PERFORMANCE_PERCENTILE = 90

try:
    from pydicom.datadict import keyword_for_tag as _pydicom_keyword_for_tag
except Exception:
    _pydicom_keyword_for_tag = None


def _generate_script_access_code(ts_ms=None):
    if ts_ms is None:
        ts_ms = int(datetime.now().timestamp() * 1000)
    seed = (int(ts_ms) // 1200000) + 171553

    # Java Random exakt nachbilden (48-bit LCG), da Python random andere Sequenz liefert.
    class _JavaRandom:
        def __init__(self, s):
            self.mult = 0x5DEECE66D
            self.add = 0xB
            self.mask = (1 << 48) - 1
            self.seed = (s ^ self.mult) & self.mask

        def _next(self, bits):
            self.seed = (self.seed * self.mult + self.add) & self.mask
            return self.seed >> (48 - bits)

        def next_int(self):
            v = self._next(32)
            return v if v < (1 << 31) else v - (1 << 32)

        def next_bytes(self, arr):
            i = 0
            ln = len(arr)
            while i < ln:
                rnd = self.next_int()
                n = min(ln - i, 4)
                for _ in range(n):
                    arr[i] = rnd & 0xFF
                    i += 1
                    rnd >>= 8

    r = _JavaRandom(seed)
    data = bytearray(10)
    r.next_bytes(data)
    crc = zlib.crc32(data) & 0xFFFFFFFF
    digits = "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    if crc == 0:
        return "0"
    out = []
    n = crc
    while n > 0:
        n, rem = divmod(n, 36)
        out.append(digits[rem])
    return "".join(reversed(out))

WORKFLOW_PROP_LABELS = {
    "1048592": "Patient Name",
    "1048608": "Patient ID",
    "1048624": "Patient Birth Date",
    "1048640": "Patient Sex",
    "3280992": "Study Description",
    "4194305": "Accession Number",
    "4194306": "Study Date",
    "4194336": "Referring Physician",
    "524368": "Series Description",
    "524384": "Modality",
    "524385": "Series Number",
}

DICOM_TAG_LABELS = {
    "1048592": "PatientName [0010,0010]",
    "1048608": "PatientID [0010,0020]",
    "3280947": "Requesting Service [0032,1033]",
    "4390932": "Private Tag [0043,0014]",
    "524320": "StudyDate [0008,0020]",
    "524336": "StudyTime [0008,0030]",
    "524368": "AccessionNumber [0008,0050]",
    "524384": "Modality [0008,0060]",
    "524385": "ModalitiesInStudy [0008,0061]",
    "524416": "InstitutionName [0008,0080]",
    "528400": "StationName [0008,1010]",
    "528432": "StudyDescription [0008,1030]",
    "528446": "SeriesDescription [0008,103E]",
}

PRIVATE_DICOM_TAG_LABELS = {
    "FFFF0002": "Study Load Order",
    "FFFF0003": "Number of Images in Series",
    "FFFF0006": "Number of Images in Display Set",
    "FFFF0008": "Display Set Orientation",
    "FFFF0009": "Display Set Type",
    "FFFF000A": "Study Load Order Reverse",
    "FFFF000B": "Number of available priors",
    "FFFF000C": "Display Set Properties",
    "FFFF000D": "Number Of Display Sets In Study",
    "FFFF000E": "Study Age",
    "FFFF000F": "Number Of 4D Runs",
    "FFFF0010": "Source Archive",
    "FFFF0011": "Study Grouped",
    "FFFF0014": "Image Stack Position",
    "FFFF0015": "Has Varying Orientation",
}


def _dicom_tag_to_hex_label(tag_value: str) -> str:
    t = (tag_value or "").strip()
    if not t:
        return ""
    known = DICOM_TAG_LABELS.get(t)
    if known:
        return known
    try:
        n = int(t)
        if n < 0:
            n = (n + (1 << 32)) & 0xFFFFFFFF
        group = (n >> 16) & 0xFFFF
        element = n & 0xFFFF
        hex_tag = f"[{group:04X},{element:04X}]"
        private_key = f"{group:04X}{element:04X}"
        private_label = PRIVATE_DICOM_TAG_LABELS.get(private_key)
        if private_label:
            return f"{private_label} {hex_tag}"
        if _pydicom_keyword_for_tag is not None:
            try:
                kw = (_pydicom_keyword_for_tag(n) or "").strip()
                if kw:
                    return f"{kw} {hex_tag}"
            except Exception:
                pass
        return hex_tag
    except Exception:
        return t


def _split_dicom_label(label: str):
    txt = (label or "").strip()
    m = re.match(r"^(.*?)\s*(\[[0-9A-Fa-f]{4},[0-9A-Fa-f]{4}\])$", txt)
    if m:
        desc = (m.group(1) or "").strip()
        num = (m.group(2) or "").strip()
        return desc, num
    return txt, ""


@app.after_request
def _set_security_headers(resp):
    resp.headers.setdefault("X-Content-Type-Options", "nosniff")
    resp.headers.setdefault("X-Frame-Options", "SAMEORIGIN")
    resp.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
    resp.headers.setdefault(
        "Content-Security-Policy",
        "default-src 'self'; img-src 'self' data:; style-src 'self' 'unsafe-inline'; script-src 'self' 'unsafe-inline'; connect-src 'self'; frame-ancestors 'self'; base-uri 'self'; form-action 'self'",
    )
    return resp


@app.errorhandler(413)
def _file_too_large(_e):
    return render_template(
        "index.html",
        error="Datei zu groß. Bitte eine Datei bis maximal 200 MB hochladen.",
    ), 413

_LAST_HANGING_PROTOCOLS_BY_ID = {}

SCREENSHOT_PERMISSION_MAPPING = {
    "Arbeitsliste bearbeiten (Geplante)": "listtext.permissions.Worklists/ScheduledWorklists/EditWorklists",
    "Erstellen (Geplante)": "listtext.permissions.Worklists/ScheduledWorklists/CreateWorklists",
    "Löschen (Geplante)": "listtext.permissions.Worklists/ScheduledWorklists/DeleteWorklists",
    "Gesperrte Arbeitslisten bearbeiten (Geplante)": "listtext.permissions.Worklists/ScheduledWorklists/EditWorklists/EditWorklistData",
    "Einträge sehen (Demo)": "listtext.permissions.Worklists/Demo/ViewWorklists",
    "Durchführen, Einträge verschieben (Demo)": "listtext.permissions.Worklists/Demo/ProcessWorklists",
    "Arbeitsliste bearbeiten (Demo)": "listtext.permissions.Worklists/Demo/ProcessWorklists",
    "Datenfilter bearbeiten (Demo)": "listtext.permissions.Worklists/DataFilter",
    "Arbeitsliste bearbeiten (Standard)": "listtext.permissions.Worklists/StandardWorklists/EditWorklists",
    "Erstellen (Standard)": "listtext.permissions.Worklists/StandardWorklists/CreateWorklists",
    "Löschen (Standard)": "listtext.permissions.Worklists/StandardWorklists/DeleteWorklists",
    "Arbeitsliste bearbeiten (Suche)": "listtext.permissions.Worklists/SearchWorklists/EditWorklists",
    "Eintrag bearbeiten (Suche)": "listtext.permissions.Worklists/SearchWorklists/EditWorklists/EditWorklistData",
    "Erstellen (Benutzer)": "listtext.permissions.Worklists/UserWorklists/CreateWorklists",
    "Arbeitsliste bearbeiten (Benutzer)": "listtext.permissions.Worklists/UserWorklists/EditWorklists",
    "Löschen (Benutzer)": "listtext.permissions.Worklists/UserWorklists/DeleteWorklists",
    "Exportieren": "listtext.permissions.Export/ExportFiles",
    "Arbeitslisten auswählen": "listtext.permissions.Worklists",
    "Drucken der Arbeitslisten": "listtext.permissions.Print/PrintWorkList",
    "Drucken der Befunde": "listtext.permissions.Print/PrintReports",
    "DICOM-Objekte vom Server löschen": "listtext.permissions.Data/Server/DeleteDicomObjects",
    "Teleradiologie": "listtext.permissions.ImageArea/Teleradiology",
    "Hintergrundaufgaben anzeigen": "listtext.permissions.JobControl",
    "Hintergrundaufgaben abbrechen": "listtext.permissions.JobControl/DeleteJob",
    "Auditing & Logging": "listtext.permissions.AuditRepository",
    "Unvalidierte Befunde ansehen": "listtext.permissions.ViewUnverifiedReports",
    "Benutzer kann Vorstudien ein-/ausschalten": "listtext.permissions.ShowPrestudyButton",
    "Notfallzugang": "listtext.permissions.EmergencyAccess",
    "Speichern von Benutzereinstellungen": "listtext.permissions.StoreUserConfiguration",
}

MANUAL_PERMISSION_TRANSLATIONS = {
    "Export": "Export",
    "Import": "Import",
    "Comments": "Kommentare",
    "ScriptWizard": "Skriptassistent",
    "FromCD": "Von CD",
    "CustomPath": "Benutzerdefinierter Pfad",
    "QBE": "QBE-Suche",
    "MediaCreationServer": "Medienerstellungsserver",
    "TeachingFile": "Lehrdatei",
    "DicomSend": "DICOM-Senden",
    "ExportNonAnnonymized": "Nicht anonymisiert exportieren",
    "ExportFiles": "Dateien exportieren",
    "BurnCD": "CD brennen",
    "ExportImages": "Bilder exportieren",
    "ExportPowerpoint": "PowerPoint exportieren",
    "EditComment": "Kommentar bearbeiten",
    "EditFavourite": "Favorit bearbeiten",
    "Email": "E-Mail",
    "IntelliSearch": "IntelliSearch",
    "XDSI": "XDS-I",
    "EditAtAGlance": "Auf einen Blick bearbeiten",
    "Profiles": "Profile",
    "TFExport": "Lehrdatei-Export",
    "CanEditImageAreaConfigHangman": "Aufhängungen konfigurieren",
    "ExpertConfigHangman": "Expertenkonfiguration Aufhängungen",
    "CanEditImageAreaConfigAnimation": "Animation konfigurieren",
}

PERMISSION_PATH_OVERRIDES = {
    "listtext.permissions.Worklists": "Client / Arbeitslisten / Arbeitslisten auswählen",
    "listtext.permissions.Worklists/Configure": "Client / Arbeitslisten / Benutzer / Änderung von Spalten und Sortierung",
    "listtext.permissions.Worklists/ScheduledWorklists/EditWorklists": "Client / Arbeitslisten / Geplante / Arbeitsliste bearbeiten",
    "listtext.permissions.Worklists/ScheduledWorklists/EditWorklists/AddWorklistEntries": "Client / Arbeitslisten / Geplante / Arbeitsliste bearbeiten / Eintrag hinzufügen",
    "listtext.permissions.Worklists/ScheduledWorklists/EditWorklists/RemoveWorklistEntries": "Client / Arbeitslisten / Geplante / Arbeitsliste bearbeiten / Eintrag löschen",
    "listtext.permissions.Worklists/ScheduledWorklists/EditWorklists/EditScheduling": "Client / Arbeitslisten / Geplante / Arbeitsliste bearbeiten / Terminplanung bearbeiten",
    "listtext.permissions.Worklists/ScheduledWorklists/EditWorklists/EditWorklistData": "Client / Arbeitslisten / Geplante / Arbeitsliste bearbeiten / Eintrag bearbeiten",
    "listtext.permissions.Worklists/ScheduledWorklists/CreateWorklists": "Client / Arbeitslisten / Geplante / Erstellen",
    "listtext.permissions.Worklists/ScheduledWorklists/DeleteWorklists": "Client / Arbeitslisten / Geplante / Löschen",
    "listtext.permissions.Worklists/ScheduledWorklists/ProcessWorklists": "Client / Arbeitslisten / Geplante / Gesperrte Arbeitslisten bearbeiten",
    "listtext.permissions.Worklists/Demo/ViewWorklists": "Client / Arbeitslisten / Demo / Einträge sehen",
    "listtext.permissions.Worklists/Demo/EditLockedWorklists": "Client / Arbeitslisten / Demo / Gesperrte Arbeitslisten bearbeiten",
    "listtext.permissions.Worklists/Demo/ProcessWorklists": "Client / Arbeitslisten / Demo / Durchführen, Einträge verschieben",
    "listtext.permissions.Worklists/Demo/EditWorklists": "Client / Arbeitslisten / Demo / Arbeitsliste bearbeiten",
    "listtext.permissions.Worklists/DataFilter": "Client / Arbeitslisten / Demo / Datenfilter bearbeiten",
    "listtext.permissions.Worklists/StandardWorklists/EditWorklists": "Client / Arbeitslisten / Standard / Arbeitsliste bearbeiten",
    "listtext.permissions.Worklists/StandardWorklists/EditWorklists/EditWorklistData": "Client / Arbeitslisten / Standard / Arbeitsliste bearbeiten / Eintrag bearbeiten",
    "listtext.permissions.Worklists/StandardWorklists/CreateWorklists": "Client / Arbeitslisten / Standard / Erstellen",
    "listtext.permissions.Worklists/StandardWorklists/DeleteWorklists": "Client / Arbeitslisten / Standard / Löschen",
    "listtext.permissions.Worklists/SearchWorklists/EditWorklists": "Client / Arbeitslisten / Suche / Arbeitsliste bearbeiten",
    "listtext.permissions.Worklists/SearchWorklists/EditWorklists/EditWorklistData": "Client / Arbeitslisten / Suche / Eintrag bearbeiten",
    "listtext.permissions.Worklists/UserWorklists/CreateWorklists": "Client / Arbeitslisten / Benutzer / Erstellen",
    "listtext.permissions.Worklists/UserWorklists/EditWorklists": "Client / Arbeitslisten / Benutzer / Arbeitsliste bearbeiten",
    "listtext.permissions.Worklists/UserWorklists/EditWorklists/RemoveWorklistEntries": "Client / Arbeitslisten / Benutzer / Arbeitsliste bearbeiten / Eintrag löschen",
    "listtext.permissions.Worklists/UserWorklists/EditWorklists/AddWorklistEntries": "Client / Arbeitslisten / Benutzer / Arbeitsliste bearbeiten / Eintrag hinzufügen",
    "listtext.permissions.Worklists/UserWorklists/EditWorklists/EditWorklistData": "Client / Arbeitslisten / Benutzer / Arbeitsliste bearbeiten / Eintrag bearbeiten",
    "listtext.permissions.Worklists/UserWorklists/DeleteWorklists": "Client / Arbeitslisten / Benutzer / Löschen",
    "listtext.permissions.Worklists/UserWorklists/DeleteWorklists/DeleteOthers": "Client / Arbeitslisten / Benutzer / Löschen / Löschen der persönlichen Arbeitslisten anderer Benutzer",
    "listtext.permissions.Worklists/UserWorklists/ExportWorklists": "Client / Arbeitslisten / Benutzer / Exportieren",
    "listtext.permissions.ImageArea/AllowToEditShortCuts": "Client / Anzeigebereich / KI / Bearbeiten von Shortcuts",
    "listtext.permissions.ImageArea/AllowToEditWindowingTree": "Client / Anzeigebereich / KI / Bearbeiten der Standardfensterwerte",
    "listtext.permissions.ImageArea/AllowedToSavePresentationStates": "Client / Anzeigebereich / KI / Speichern von Presentation States",
    "listtext.permissions.ImageArea/AllowedToDoManualSplitAndSort": "Client / Anzeigebereich / KI / Manuelles Aufteilen und Sortieren",
    "listtext.permissions.ImageArea/AllowedToEditHangingProtocols": "Client / Anzeigebereich / KI / Erstellen, Ändern und Speichern von Aufhängungen",
    "listtext.permissions.ImageArea/ScreenSharing": "Client / Anzeigebereich / KI / Live View",
    "listtext.permissions.ImageArea/AllowedToSeeKeyInDescription": "Client / Anzeigebereich / KI / Konfigurationsschlüssel in Beschreibungen",
    "listtext.permissions.ImageArea/ConfigureTools": "Client / Anzeigebereich / KI / Werkzeuge konfigurieren",
    "listtext.permissions.ImageArea/UseSmartHangings": "Client / Anzeigebereich / KI / Smart Hanging Protocol trainieren/zurücksetzen/freigeben",
    "listtext.permissions.ImageArea/AllowedToCreateAndSaveSessions": "Client / Anzeigebereich / KI / Erstellen und Speichern von Sessions",
    "listtext.permissions.XDSI": "Client / XDSI",
    "listtext.permissions.XDSI/ChangeResultTable": "Client / XDSI / Benutzer kann Sortierung der Liste anpassen",

    "listtext.permissions.Configuration": "Client / Konfiguration",
    "listtext.permissions.Configuration/ImageArea": "Client / Konfiguration / Anzeigebereich",
    "listtext.permissions.Configuration/ImageArea/CanEditImageAreaConfigDemographics": "Client / Konfiguration / Anzeigebereich / Bildbeschriftungen",
    "listtext.permissions.Configuration/ImageArea/CanEditImageAreaConfigWindowingPresets": "Client / Konfiguration / Anzeigebereich / Standardfensterwerte",
    "listtext.permissions.Configuration/ImageArea/HangingProtocols": "Client / Konfiguration / Anzeigebereich / Aufhängungen",
    "listtext.permissions.Configuration/ImageArea/HangingProtocols/ConditionSets": "Client / Konfiguration / Anzeigebereich / Aufhängungen / Bedingungssets",
    "listtext.permissions.Configuration/ImageArea/HangingProtocols/SortOrders": "Client / Konfiguration / Anzeigebereich / Aufhängungen / Sortierreihenfolgen",
    "listtext.permissions.Configuration/ImageArea/HangingProtocols/TagsCFind": "Client / Konfiguration / Anzeigebereich / Aufhängungen / Felder C-Find",
    "listtext.permissions.Configuration/ImageArea/HangingProtocols/HangingProtocols": "Client / Konfiguration / Anzeigebereich / Aufhängungen / Aufhängungen",
    "listtext.permissions.Configuration/ImageArea/HangingProtocols/SplitAndSort": "Client / Konfiguration / Anzeigebereich / Aufhängungen / Aufteilen und Sortieren",
    "listtext.permissions.Configuration/ImageArea/HangingProtocols/DefaultHangingProtocol": "Client / Konfiguration / Anzeigebereich / Aufhängungen / Standard-Aufhängung",
    "listtext.permissions.Configuration/ImageArea/CanEditImageAreaConfigColorsFonts": "Client / Konfiguration / Anzeigebereich / Markups",
    "listtext.permissions.Configuration/ImageArea/CanEditImageAreaConfigDisplay": "Client / Konfiguration / Anzeigebereich / Anzeigefenster",
    "listtext.permissions.Configuration/ImageArea/CanEditImageAreaConfigMonitors": "Client / Konfiguration / Anzeigebereich / Monitore",
    "listtext.permissions.Configuration/ImageArea/CanEditImageAreaConfigDisplaySetPanel": "Client / Konfiguration / Anzeigebereich / Serienpalette",
    "listtext.permissions.Configuration/ImageArea/CanEditImageAreaConfigDialogs": "Client / Konfiguration / Anzeigebereich / Dialoge",
    "listtext.permissions.Configuration/ImageArea/CanEditImageAreaConfigGeneral": "Client / Konfiguration / Anzeigebereich / Allgemeine Einstellungen",
    "listtext.permissions.Configuration/ImageArea/CanEditImageAreaConfigSessionNames": "Client / Konfiguration / Anzeigebereich / Session-Namen",

    "listtext.permissions.Configuration/Security": "Client / Konfiguration / Anzeige von Berechtigungen",
    "listtext.permissions.Configuration/Security/EditPermissions": "Client / Konfiguration / Anzeige von Berechtigungen / Berechtigungen bearbeiten",
    "listtext.permissions.Configuration/Security/Impersonate": "Client / Konfiguration / Anzeige von Berechtigungen / Als Rolle/Admin anmelden",
    "listtext.permissions.Configuration/Security/ModifyRoles": "Client / Konfiguration / Anzeige von Berechtigungen / Rollen ändern",

    "listtext.permissions.Configuration/PrinterConfiguration": "Client / Konfiguration / Druckerkonfiguration",

    "listtext.permissions.Configuration/Preferences": "Client / Konfiguration / Generelle Einstellungen",
    "listtext.permissions.Configuration/Preferences/EditDicomAuthenticationType": "Client / Konfiguration / Generelle Einstellungen / Anzeige/Ändern der verw. DICOM-Authentifizierung",
    "listtext.permissions.Configuration/Preferences/ListAreaConfiguration": "Client / Konfiguration / Generelle Einstellungen / Listenbereich-Einstellungen",
    "listtext.permissions.Configuration/Preferences/PluginsConfiguration": "Client / Konfiguration / Generelle Einstellungen / Einstellungen für Plugins",
    "listtext.permissions.Configuration/Preferences/TextAreaConfiguration": "Client / Konfiguration / Generelle Einstellungen / Textbereich-Einstellungen",
    "listtext.permissions.Configuration/Preferences/ExportConfiguration": "Client / Konfiguration / Generelle Einstellungen / Export Einstellungen",

    "listtext.permissions.Configuration/Local": "Client / Konfiguration / Lokale Konfiguration",
    "listtext.permissions.Configuration/Local/EditCategories": "Client / Konfiguration / Lokale Konfiguration / Arbeitsstationsgruppen bearbeiten",
    "listtext.permissions.Configuration/Local/LoginConfiguration": "Client / Konfiguration / Lokale Konfiguration / Anmelde-Konfiguration",

    "listtext.permissions.Configuration/Profiles": "Client / Konfiguration / Profiles",
    "listtext.permissions.Configuration/EditTextAreaLayout": "Client / Konfiguration / Textbereich-Layout bearbeiten",
    "listtext.permissions.Configuration/EditGroupConfiguration": "Client / Konfiguration / Rollen/Gruppen Konfiguration",

    "listtext.permissions.Import": "Client / Import",
    "listtext.permissions.Import/ImportFiles": "Client / Import / Textbereich-Datenimportmappe – Import",
    "listtext.permissions.Import/ImportFiles/ExportImportedFiles": "Client / Import / Textbereich-Datenimportmappe – Re-Export",

    "listtext.permissions.Print": "Client / Drucken",
    "listtext.permissions.Print/PrintWorkList": "Client / Drucken / Drucken der Arbeitslisten",
    "listtext.permissions.Print/PrintReport": "Client / Drucken / Drucken der Befunde",

    "listtext.permissions.DeleteDicomObjects": "Client / DICOM-Objekte vom Server löschen",
    "listtext.permissions.DeleteDicomObjects/DeleteKeyImages": "Client / DICOM-Objekte vom Server löschen / Key Images vom Server löschen",
    "listtext.permissions.DeleteDicomObjects/DeletePresentationStates": "Client / DICOM-Objekte vom Server löschen / Presentation States vom Server löschen",
    "listtext.permissions.DeleteDicomObjects/DeletePresentationStates/DeleteOwnPresentationStates": "Client / DICOM-Objekte vom Server löschen / Presentation States vom Server löschen / Eigene Presentation States vom Server löschen",
    "listtext.permissions.DeleteDicomObjects/DeleteSessions": "Client / DICOM-Objekte vom Server löschen / Sessions vom Server löschen",

    "listtext.permissions.Teleradiology": "Client / Teleradiologie",

    "listtext.permissions.IntegratedReporting/CancelTask": "Client / DeepUnity Discovery / Befundungsaufgabe stornieren",

    "listtext.permissions.AdvancedDicomInformation": "Client / Zusätzliche DICOM-Information im Listenbereich",

    "listtext.permissions.ScriptWizard": "Client / ScriptWizard",
    "listtext.permissions.ScriptWizard/EditScript": "Client / ScriptWizard / Skripte editieren",

    "listtext.permissions.JobControl": "Client / Hintergrundaufgaben",
    "listtext.permissions.JobControl/ShowAllJobs": "Client / Hintergrundaufgaben / Alle Hintergrundaufgaben anzeigen",
    "listtext.permissions.JobControl/DeleteJob": "Client / Hintergrundaufgaben / Abbrechen von Hintergrundaufgaben",
    "listtext.permissions.ChangeStudyStatus": "Client / Hintergrundaufgaben / Status einer Studie ändern",
    "listtext.permissions.AuditRepository": "Client / Hintergrundaufgaben / Auditing & Logging",

    "listtext.permissions.Search/Barcode": "Client / Suche / Barcode",
    "listtext.permissions.Search/Advanced": "Client / Suche / Advanced",
    "listtext.permissions.Search/QBE": "Client / Suche / QBE",
    "listtext.permissions.Search/IntelliSearch": "Client / Suche / IntelliSearch",

    "listtext.permissions.Export": "Client / Export",
    "listtext.permissions.Export/ExportClinicalTrials": "Client / Export / Forschungsserien exportieren",
    "listtext.permissions.Export/InternetStudyShare": "Client / Export / Studienfreigabe für das Internet",
    "listtext.permissions.ViewUnverifiedReports": "Client / Export / Unvalidierte Befunde ansehen",
    "listtext.permissions.ShowPrestudyButton": "Client / Export / Benutzer kann Vorstudien ein-/ausschalten",
    "listtext.permissions.AllowToSendEmailDump": "Client / Export / Senden von Fehlermeldungen via E-Mail",
    "listtext.permissions.EditKeywords": "Client / Export / Schlagwörter bearbeiten",
    "listtext.permissions.EmergencyAccess": "Client / Export / Notfallzugang",
    "listtext.permissions.EditPriors": "Client / Export / Regel für Vorstudien",
    "listtext.permissions.StoreUserConfiguration": "Client / Export / Speichern von Benutzereinstellungen",
    "listtext.permissions.ViewUnreported": "Client / Export / Unbefundete Studien ansehen",
}

_PERMISSION_LABEL_CACHE = None


def _load_permission_label_map():
    global _PERMISSION_LABEL_CACHE
    if _PERMISSION_LABEL_CACHE is not None:
        return _PERMISSION_LABEL_CACHE

    labels = {}
    plugins_dir = Path(__file__).resolve().parent.parent / "plugins"
    if plugins_dir.exists():
        for jar_path in plugins_dir.glob("*.jar"):
            try:
                with zipfile.ZipFile(jar_path) as zf:
                    for entry in zf.namelist():
                        if not (entry.endswith("_de.properties") or entry == "messages_de.properties"):
                            continue
                        try:
                            content = zf.read(entry).decode("utf-8", "ignore")
                        except Exception:
                            continue
                        for ln in content.splitlines():
                            if "=" not in ln:
                                continue
                            stripped = ln.strip()
                            if not stripped or stripped.startswith("#"):
                                continue
                            k, v = ln.split("=", 1)
                            k = k.strip()
                            v = v.strip()
                            if k and k not in labels:
                                labels[k] = v
            except Exception:
                continue

    _PERMISSION_LABEL_CACHE = labels
    return labels


def _permission_segment_to_de(segment, labels):
    if segment in MANUAL_PERMISSION_TRANSLATIONS:
        return MANUAL_PERMISSION_TRANSLATIONS[segment]
    if segment in labels:
        return labels[segment]
    for prefix in [
        "Permission.",
        "Permissions.",
        "Config.",
        "ConfigHangingProtocols.",
        "WorklistUI.",
        "WorklistLibraryIdentifier.",
    ]:
        key = f"{prefix}{segment}"
        if key in labels:
            return labels[key]
    return segment


def _permission_display_path(key, labels):
    if key in PERMISSION_PATH_OVERRIDES:
        return PERMISSION_PATH_OVERRIDES[key]

    path_part = key[len("listtext.permissions"):].lstrip("/")
    segments = [s[1:] if s.startswith(".") else s for s in path_part.split("/") if s]
    de_segments = [_permission_segment_to_de(s, labels) for s in segments]
    if not de_segments:
        return "Client"
    return "Client / " + " / ".join(de_segments)


def _normalize_item_name(name):
    return name[len("shared|."):] if name.startswith("shared|.") else name


def _parse_last_updated(value):
    if not value:
        return None
    # Beispiel: Fri Aug 08 11:02:55 CEST 2025
    cleaned = re.sub(r"\s+[A-Z]{3,4}\s+", " ", value.strip())
    try:
        return datetime.strptime(cleaned, "%a %b %d %H:%M:%S %Y")
    except Exception:
        return None


def _flatten_xml_values(node, path=""):
    lines = []
    tag = node.tag
    current = f"{path}/{tag}" if path else tag

    for k, v in sorted(node.attrib.items(), key=lambda x: x[0]):
        lines.append(f"{current}.@{k} = {v}")

    text = (node.text or "").strip()
    if text:
        lines.append(f"{current} = {text}")

    for child in list(node):
        lines.extend(_flatten_xml_values(child, current))

    return lines


def _safe_svg_filename(icon_name, fallback):
    base = (icon_name or "").strip() or fallback
    safe = re.sub(r"[^A-Za-z0-9._-]+", "_", base).strip("._-")
    if not safe:
        safe = fallback
    if not safe.lower().endswith(".svg"):
        safe += ".svg"
    return safe


def _safe_download_filename(name, fallback, extension):
    base = (name or "").strip() or fallback
    safe = re.sub(r"[^A-Za-z0-9._-]+", "_", base).strip("._-")
    if not safe:
        safe = fallback
    ext = extension if extension.startswith(".") else f".{extension}"
    if not safe.lower().endswith(ext.lower()):
        safe += ext
    return safe


def _build_agh_binary(xml_text, token):
    xml_bytes = (xml_text or "").encode("utf-8")
    sig_bytes = (token or "").encode("utf-8")

    out = bytearray()
    out.extend((1).to_bytes(4, "big"))
    out.extend(len(sig_bytes).to_bytes(4, "big"))
    out.extend(sig_bytes)
    out.extend(len(xml_bytes).to_bytes(4, "big"))
    out.extend(xml_bytes)
    return bytes(out)


def _calc_xml_signature(xml_bytes, salt_bytes):
    md5_digest = hashlib.md5(xml_bytes + bytes(salt_bytes)).digest()
    return base64.b64encode(md5_digest).decode("ascii")


def _calc_agh_signature(xml_bytes):
    # ExportParameter.Q1_2009 salt
    return _calc_xml_signature(xml_bytes, [20, 58, 63, 4])


def _calc_agd_signature(xml_bytes):
    # ExportParameter.DESCRIPTORS salt
    return _calc_xml_signature(xml_bytes, [23, 50, 12, 7])


def _build_agd_data_uri_from_descriptors(descriptor_xml_blocks):
    blocks = [b for b in (descriptor_xml_blocks or []) if (b or "").strip()]
    payload = "<descriptorsImportExport>\n"
    if blocks:
        payload += "\n".join(blocks) + "\n"
    payload += "</descriptorsImportExport>"
    payload_bytes = payload.encode("utf-8")
    signature = _calc_agd_signature(payload_bytes)
    agd_binary = _build_agh_binary(payload, token=signature)
    return "data:application/octet-stream;base64," + base64.b64encode(agd_binary).decode("ascii")


def _detect_xml_kind(root):
    raw_tag = (getattr(root, "tag", "") or "").strip()
    local_tag = raw_tag.split("}", 1)[-1] if "}" in raw_tag else raw_tag
    if local_tag.lower() == "licensefile":
        return "licensefile"
    return "exportxml"


def _description_in_parentheses(text):
    raw = (text or "").strip()
    if not raw:
        return ""
    hits = re.findall(r"\(([^()]*)\)", raw)
    cleaned = [h.strip() for h in hits if h.strip()]
    if cleaned:
        return " | ".join(cleaned)
    return raw


def _decode_performance_csv(raw):
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin1"):
        try:
            return raw.decode(enc), enc
        except Exception:
            continue
    return raw.decode("latin1", errors="replace"), "latin1"


def _parse_float_de(value):
    s = ("" if value is None else str(value)).strip()
    if not s:
        return None
    s = s.replace(".", "").replace(",", ".") if re.match(r"^-?\d{1,3}(\.\d{3})+,\d+$", s) else s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


def _parse_int_like(value):
    v = _parse_float_de(value)
    if v is None:
        return None
    try:
        return int(round(v))
    except Exception:
        return None


def _fmt_decimal(value, digits=1):
    if value is None:
        return ""
    try:
        return (f"{float(value):.{digits}f}").replace(".", ",")
    except Exception:
        return ""


def _fmt_ms_seconds(ms, digits=2):
    if ms is None:
        return ""
    try:
        return (f"{float(ms) / 1000.0:.{digits}f} s").replace(".", ",")
    except Exception:
        return ""


def _performance_percentile_label(pct):
    try:
        p = float(pct)
        return f"P{int(p)}" if p.is_integer() else f"P{str(p).replace('.', ',')}"
    except Exception:
        return "Perzentil"


def _performance_percentile_hint(pct):
    label = _performance_percentile_label(pct)
    try:
        p = float(pct)
        pct_text = f"{int(p)}" if p.is_integer() else str(p).replace('.', ',')
    except Exception:
        pct_text = str(pct)
    return (
        f"{label} = berechnetes {pct_text}. Perzentil der sortierten Messwerte; "
        "bei kleinen Gruppen oder interpolierten Werten ist das kein exakter Zähler "
        "für genau diesen Anteil der Messungen."
    )


def _percentile(values, pct):
    vals = sorted(v for v in values if v is not None)
    if not vals:
        return None
    if len(vals) == 1:
        return vals[0]
    k = (len(vals) - 1) * (pct / 100.0)
    lo = int(k)
    hi = min(lo + 1, len(vals) - 1)
    frac = k - lo
    return vals[lo] * (1 - frac) + vals[hi] * frac


def _avg(values):
    vals = [v for v in values if v is not None]
    if not vals:
        return None
    return sum(vals) / len(vals)


def _parse_performance_date(value):
    raw = (value or "").strip()
    for fmt in ("%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(raw, fmt)
        except Exception:
            pass
    return None


def analyze_performance_csv(file_storage, filename="performance.csv"):
    raw = file_storage.read() if hasattr(file_storage, "read") else bytes(file_storage or b"")
    text, encoding = _decode_performance_csv(raw)
    lines = [ln for ln in text.splitlines() if ln.strip()]
    if not lines:
        raise ValueError("Leere CSV-Datei")

    delimiter = ";"
    try:
        sample = "\n".join(lines[:20])
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
        delimiter = dialect.delimiter
    except Exception:
        delimiter = ";" if lines[0].count(";") >= lines[0].count(",") else ","

    reader = csv.DictReader(lines, delimiter=delimiter)
    raw_rows = [r for r in reader if any((v or "").strip() for v in r.values())]
    headers = reader.fieldnames or []

    expected = [
        "Start Date", "Study Instance UID", "First Display", "Full Study", "CFind Duration",
        "Modality", "Object count", "Prior", "Hanging", "Base Study", "Bandwidth", "Hangup Type", "User"
    ]
    missing = [h for h in expected if h not in headers]

    rows = []
    for idx, r in enumerate(raw_rows, start=1):
        start_raw = (r.get("Start Date") or "").strip()
        start_dt = _parse_performance_date(start_raw)
        first_ms = _parse_float_de(r.get("First Display"))
        full_ms = _parse_float_de(r.get("Full Study"))
        cfind_ms = _parse_float_de(r.get("CFind Duration"))
        object_count = _parse_int_like(r.get("Object count"))
        bandwidth = _parse_float_de(r.get("Bandwidth"))
        modality = (r.get("Modality") or "").strip()
        prior = (r.get("Prior") or "").strip()
        hanging = (r.get("Hanging") or "").strip()
        hangup_type = (r.get("Hangup Type") or "").strip()
        user = (r.get("User") or "").strip()
        study_uid = (r.get("Study Instance UID") or "").strip()
        base_study = (r.get("Base Study") or "").strip()
        slow_first = first_ms is not None and first_ms >= 10000
        slow_full = full_ms is not None and full_ms >= 30000
        search_text = " ".join([
            start_raw, study_uid, modality, prior, hanging, base_study, hangup_type, user,
            str(r.get("Object count") or ""), str(r.get("Bandwidth") or ""),
        ]).lower()
        rows.append({
            "rowNr": idx,
            "startDate": start_raw,
            "startSort": start_dt.isoformat() if start_dt else start_raw,
            "startEpochMs": int(start_dt.timestamp() * 1000) if start_dt else None,
            "studyInstanceUid": study_uid,
            "firstDisplayMs": first_ms,
            "firstDisplayRaw": (r.get("First Display") or "").strip(),
            "firstDisplaySec": _fmt_ms_seconds(first_ms),
            "fullStudyMs": full_ms,
            "fullStudyRaw": (r.get("Full Study") or "").strip(),
            "fullStudySec": _fmt_ms_seconds(full_ms),
            "cfindMs": cfind_ms,
            "cfindRaw": (r.get("CFind Duration") or "").strip(),
            "cfindSec": _fmt_ms_seconds(cfind_ms),
            "modality": modality,
            "objectCount": object_count if object_count is not None else (r.get("Object count") or "").strip(),
            "prior": prior,
            "hanging": hanging,
            "baseStudy": base_study,
            "bandwidth": _fmt_decimal(bandwidth, 2) if bandwidth is not None else (r.get("Bandwidth") or "").strip(),
            "bandwidthValue": bandwidth,
            "hangupType": hangup_type,
            "user": user,
            "isSlowFirst": slow_first,
            "isSlowFull": slow_full,
            "isSlow": slow_first or slow_full,
            "searchText": search_text,
        })

    rows.sort(key=lambda x: (x.get("startSort") or "", x.get("rowNr") or 0))

    first_values = [r["firstDisplayMs"] for r in rows if r.get("firstDisplayMs") is not None]
    full_values = [r["fullStudyMs"] for r in rows if r.get("fullStudyMs") is not None]
    cfind_values = [r["cfindMs"] for r in rows if r.get("cfindMs") is not None]
    bandwidth_values = [r["bandwidthValue"] for r in rows if r.get("bandwidthValue") is not None]
    users = sorted({r["user"] for r in rows if r.get("user")})
    modalities = sorted({r["modality"] for r in rows if r.get("modality")})
    hanging_names = sorted({r["hanging"] for r in rows if r.get("hanging")})
    hangup_types = sorted({r["hangupType"] for r in rows if r.get("hangupType")})
    dates = [_parse_performance_date(r["startDate"]) for r in rows if r.get("startDate")]
    dates = [d for d in dates if d is not None]

    def _stats_for_group(group_rows, key, label):
        fvals = [r["firstDisplayMs"] for r in group_rows if r.get("firstDisplayMs") is not None]
        fulls = [r["fullStudyMs"] for r in group_rows if r.get("fullStudyMs") is not None]
        object_counts = [r["objectCount"] for r in group_rows if isinstance(r.get("objectCount"), int)]
        return {
            key: label,
            "count": len(group_rows),
            "objectCountAvg": _fmt_decimal(_avg(object_counts), 1),
            "firstAvg": _fmt_ms_seconds(_avg(fvals)),
            "firstP90": _fmt_ms_seconds(_percentile(fvals, PERFORMANCE_PERCENTILE)),
            "fullAvg": _fmt_ms_seconds(_avg(fulls)),
            "fullP90": _fmt_ms_seconds(_percentile(fulls, PERFORMANCE_PERCENTILE)),
            "notHungCount": sum(1 for r in group_rows if r.get("hangupType") == "NOT_HUNG_UP_INITIALLY"),
            "priorCount": sum(1 for r in group_rows if (r.get("prior") or "").lower() == "true"),
            "slowCount": sum(1 for r in group_rows if r.get("isSlow")),
        }

    modality_groups = []
    by_modality = defaultdict(list)
    for r in rows:
        by_modality[r.get("modality") or "-"] .append(r)
    for label, group_rows in by_modality.items():
        modality_groups.append(_stats_for_group(group_rows, "modality", label))
    modality_groups.sort(key=lambda x: (-x["count"], x["modality"]))

    hanging_groups = []
    by_hanging = defaultdict(list)
    for r in rows:
        by_hanging[r.get("hanging") or "Ohne Hanging"].append(r)
    for label, group_rows in by_hanging.items():
        hanging_groups.append(_stats_for_group(group_rows, "hanging", label))
    hanging_groups.sort(key=lambda x: (-x["count"], x["hanging"]))

    user_groups = []
    by_user = defaultdict(list)
    for r in rows:
        by_user[r.get("user") or "-"] .append(r)
    for label, group_rows in by_user.items():
        user_groups.append(_stats_for_group(group_rows, "user", label))
    user_groups.sort(key=lambda x: (-x["count"], x["user"]))

    hangup_counts = Counter(r.get("hangupType") or "-" for r in rows)

    summary = {
        "rows_total": len(rows),
        "users_total": len(users),
        "modalities_total": len(modalities),
        "hanging_total": len(hanging_names),
        "prior_count": sum(1 for r in rows if (r.get("prior") or "").lower() == "true"),
        "not_hung_count": sum(1 for r in rows if r.get("hangupType") == "NOT_HUNG_UP_INITIALLY"),
        "initial_hung_count": sum(1 for r in rows if r.get("hangupType") == "HUNG_UP_BY_INITIAL_HP"),
        "slow_count": sum(1 for r in rows if r.get("isSlow")),
        "slow_first_count": sum(1 for r in rows if r.get("isSlowFirst")),
        "slow_full_count": sum(1 for r in rows if r.get("isSlowFull")),
        "first_avg": _fmt_ms_seconds(_avg(first_values)),
        "first_p50": _fmt_ms_seconds(_percentile(first_values, 50)),
        "first_p90": _fmt_ms_seconds(_percentile(first_values, PERFORMANCE_PERCENTILE)),
        "full_avg": _fmt_ms_seconds(_avg(full_values)),
        "full_p50": _fmt_ms_seconds(_percentile(full_values, 50)),
        "full_p90": _fmt_ms_seconds(_percentile(full_values, PERFORMANCE_PERCENTILE)),
        "cfind_avg": _fmt_ms_seconds(_avg(cfind_values)),
        "bandwidth_avg": _fmt_decimal(_avg(bandwidth_values), 2),
        "date_min": min(dates).strftime("%d.%m.%Y %H:%M:%S") if dates else "",
        "date_max": max(dates).strftime("%d.%m.%Y %H:%M:%S") if dates else "",
    }

    return {
        "xml_kind": "performancecsv",
        "root": "performance.csv",
        "filename": filename,
        "encoding": encoding,
        "delimiter": delimiter,
        "headers": headers,
        "missing_headers": missing,
        "performance_summary": summary,
        "performance_percentile": PERFORMANCE_PERCENTILE,
        "performance_percentile_label": _performance_percentile_label(PERFORMANCE_PERCENTILE),
        "performance_percentile_hint": _performance_percentile_hint(PERFORMANCE_PERCENTILE),
        "performance_rows": rows,
        "performance_modalities": modality_groups,
        "performance_hangings": hanging_groups,
        "performance_users": user_groups,
        "performance_filters": {
            "users": users,
            "modalities": modalities,
            "hangup_types": hangup_types,
        },
        "performance_hangup_counts": [
            {"hangupType": k, "count": v}
            for k, v in sorted(hangup_counts.items(), key=lambda x: (-x[1], x[0]))
        ],
    }


def _fmt_delta_seconds(delta_ms, digits=2):
    if delta_ms is None:
        return ""
    try:
        value = float(delta_ms) / 1000.0
        sign = "+" if value > 0 else ""
        return (f"{sign}{value:.{digits}f} s").replace(".", ",")
    except Exception:
        return ""


def _fmt_delta_decimal(delta, digits=1):
    if delta is None:
        return ""
    try:
        value = float(delta)
        sign = "+" if value > 0 else ""
        return (f"{sign}{value:.{digits}f}").replace(".", ",")
    except Exception:
        return ""


def _fmt_delta_int(delta):
    if delta is None:
        return ""
    try:
        value = int(delta)
        return f"+{value}" if value > 0 else str(value)
    except Exception:
        return ""


def _filename_stem(filename):
    name = Path(filename or "").name.strip()
    if not name:
        return ""
    lower = name.lower()
    for suffix in (".csv", ".xml"):
        if lower.endswith(suffix):
            return name[: -len(suffix)]
    return name


def _split_filename_parts(stem):
    # Keep only meaningful parts, independent of whether the difference is at
    # the beginning, in the middle or at the end before the file extension.
    return [p for p in re.split(r"[-_.\s]+", stem or "") if p]


def _common_prefix_len(a, b):
    n = min(len(a), len(b))
    i = 0
    while i < n and str(a[i]).lower() == str(b[i]).lower():
        i += 1
    return i


def _common_suffix_len(a, b, prefix_len=0):
    max_len = min(len(a), len(b)) - prefix_len
    i = 0
    while i < max_len and str(a[len(a) - 1 - i]).lower() == str(b[len(b) - 1 - i]).lower():
        i += 1
    return i


def _pretty_compare_label(text, fallback):
    raw = (text or "").strip(" _-.()[]{}")
    if not raw:
        raw = fallback
    raw = re.sub(r"[-_.]+", " ", raw).strip()
    raw = re.sub(r"\s+", " ", raw)
    if not raw:
        raw = fallback
    return raw[:60] + "…" if len(raw) > 60 else raw


def _build_filename_compare_labels(filename_a, filename_b):
    stem_a = _filename_stem(filename_a)
    stem_b = _filename_stem(filename_b)
    if not stem_a or not stem_b or stem_a.lower() == stem_b.lower():
        return (_pretty_compare_label(stem_a or filename_a or "Datei 1", "Datei 1"),
                _pretty_compare_label(stem_b or filename_b or "Datei 2", "Datei 2"))

    # Prefer delimiter-aware comparison first. This produces clean labels for:
    # - beginning: baseline_performance.csv vs after_performance.csv
    # - middle:    performance_baseline_2026.csv vs performance_after_2026.csv
    # - end:       performance_2026_baseline.csv vs performance_2026_after.csv
    parts_a = _split_filename_parts(stem_a)
    parts_b = _split_filename_parts(stem_b)
    if parts_a and parts_b and (len(parts_a) > 1 or len(parts_b) > 1):
        prefix_parts = _common_prefix_len(parts_a, parts_b)
        suffix_parts = _common_suffix_len(parts_a, parts_b, prefix_parts)
        end_a = len(parts_a) - suffix_parts if suffix_parts else len(parts_a)
        end_b = len(parts_b) - suffix_parts if suffix_parts else len(parts_b)
        diff_parts_a = parts_a[prefix_parts:end_a]
        diff_parts_b = parts_b[prefix_parts:end_b]
        if diff_parts_a and diff_parts_b:
            label_a = _pretty_compare_label(" ".join(diff_parts_a), stem_a)
            label_b = _pretty_compare_label(" ".join(diff_parts_b), stem_b)
            if label_a.lower() != label_b.lower():
                return label_a, label_b

    # Fallback for filenames without separators, e.g. perfbaseline.csv vs perfafter.csv.
    prefix_len = _common_prefix_len(stem_a, stem_b)
    suffix_len = _common_suffix_len(stem_a, stem_b, prefix_len)
    end_a = len(stem_a) - suffix_len if suffix_len else len(stem_a)
    end_b = len(stem_b) - suffix_len if suffix_len else len(stem_b)
    diff_a = stem_a[prefix_len:end_a]
    diff_b = stem_b[prefix_len:end_b]

    label_a = _pretty_compare_label(diff_a, stem_a)
    label_b = _pretty_compare_label(diff_b, stem_b)

    if label_a.lower() == label_b.lower():
        label_a = _pretty_compare_label(stem_a, "Datei 1")
        label_b = _pretty_compare_label(stem_b, "Datei 2")

    return label_a, label_b


def _performance_numeric_stats(group_rows):
    fvals = [r["firstDisplayMs"] for r in group_rows if r.get("firstDisplayMs") is not None]
    fulls = [r["fullStudyMs"] for r in group_rows if r.get("fullStudyMs") is not None]
    object_counts = [r["objectCount"] for r in group_rows if isinstance(r.get("objectCount"), int)]
    return {
        "count": len(group_rows),
        "objectAvgValue": _avg(object_counts),
        "objectAvg": _fmt_decimal(_avg(object_counts), 1),
        "firstAvgValue": _avg(fvals),
        "firstAvg": _fmt_ms_seconds(_avg(fvals)),
        "firstP90Value": _percentile(fvals, PERFORMANCE_PERCENTILE),
        "firstP90": _fmt_ms_seconds(_percentile(fvals, PERFORMANCE_PERCENTILE)),
        "fullAvgValue": _avg(fulls),
        "fullAvg": _fmt_ms_seconds(_avg(fulls)),
        "fullP90Value": _percentile(fulls, PERFORMANCE_PERCENTILE),
        "fullP90": _fmt_ms_seconds(_percentile(fulls, PERFORMANCE_PERCENTILE)),
        "slowCount": sum(1 for r in group_rows if r.get("isSlow")),
    }



def _build_performance_comparison(result_a, result_b, filename_a, filename_b):
    rows_a = result_a.get("performance_rows") or []
    rows_b = result_b.get("performance_rows") or []
    label_a, label_b = _build_filename_compare_labels(filename_a, filename_b)

    def group_by(rows, key, fallback="-"):
        out = {}
        for r in rows:
            label = r.get(key) or fallback
            out.setdefault(label, []).append(r)
        return out

    def compare_groups(key, fallback):
        grouped_a = group_by(rows_a, key, fallback)
        grouped_b = group_by(rows_b, key, fallback)
        labels = sorted(set(grouped_a.keys()) | set(grouped_b.keys()), key=lambda x: x.lower())
        out = []
        for label in labels:
            a = _performance_numeric_stats(grouped_a.get(label, []))
            b = _performance_numeric_stats(grouped_b.get(label, []))
            out.append({
                "label": label,
                "countA": a["count"],
                "countB": b["count"],
                "countDelta": _fmt_delta_int(b["count"] - a["count"]),
                "objectAvgA": a["objectAvg"] or "-",
                "objectAvgB": b["objectAvg"] or "-",
                "objectAvgDelta": _fmt_delta_decimal(None if a["objectAvgValue"] is None or b["objectAvgValue"] is None else b["objectAvgValue"] - a["objectAvgValue"], 1),
                "firstAvgA": a["firstAvg"] or "-",
                "firstAvgB": b["firstAvg"] or "-",
                "firstAvgMsA": a["firstAvgValue"],
                "firstAvgMsB": b["firstAvgValue"],
                "firstAvgDelta": _fmt_delta_seconds(None if a["firstAvgValue"] is None or b["firstAvgValue"] is None else b["firstAvgValue"] - a["firstAvgValue"]),
                "firstP90A": a["firstP90"] or "-",
                "firstP90B": b["firstP90"] or "-",
                "firstP90MsA": a["firstP90Value"],
                "firstP90MsB": b["firstP90Value"],
                "firstP90Delta": _fmt_delta_seconds(None if a["firstP90Value"] is None or b["firstP90Value"] is None else b["firstP90Value"] - a["firstP90Value"]),
                "fullAvgA": a["fullAvg"] or "-",
                "fullAvgB": b["fullAvg"] or "-",
                "fullAvgMsA": a["fullAvgValue"],
                "fullAvgMsB": b["fullAvgValue"],
                "fullAvgDelta": _fmt_delta_seconds(None if a["fullAvgValue"] is None or b["fullAvgValue"] is None else b["fullAvgValue"] - a["fullAvgValue"]),
                "fullP90A": a["fullP90"] or "-",
                "fullP90B": b["fullP90"] or "-",
                "fullP90MsA": a["fullP90Value"],
                "fullP90MsB": b["fullP90Value"],
                "fullP90Delta": _fmt_delta_seconds(None if a["fullP90Value"] is None or b["fullP90Value"] is None else b["fullP90Value"] - a["fullP90Value"]),
                "slowA": a["slowCount"],
                "slowB": b["slowCount"],
                "slowDelta": _fmt_delta_int(b["slowCount"] - a["slowCount"]),
            })
        out.sort(key=lambda r: (-(r["countA"] + r["countB"]), r["label"].lower()))
        return out

    stats_a = _performance_numeric_stats(rows_a)
    stats_b = _performance_numeric_stats(rows_b)

    def delta_ms(a, b):
        return _fmt_delta_seconds(None if a is None or b is None else b - a)

    summary = {
        "filename_a": filename_a,
        "filename_b": filename_b,
        "label_a": label_a,
        "label_b": label_b,
        "rows_a": len(rows_a),
        "rows_b": len(rows_b),
        "rows_delta": _fmt_delta_int(len(rows_b) - len(rows_a)),
        "modalities_a": result_a.get("performance_summary", {}).get("modalities_total", 0),
        "modalities_b": result_b.get("performance_summary", {}).get("modalities_total", 0),
        "users_a": result_a.get("performance_summary", {}).get("users_total", 0),
        "users_b": result_b.get("performance_summary", {}).get("users_total", 0),
        "slow_a": stats_a["slowCount"],
        "slow_b": stats_b["slowCount"],
        "slow_delta": _fmt_delta_int(stats_b["slowCount"] - stats_a["slowCount"]),
        "first_avg_a": stats_a["firstAvg"] or "-",
        "first_avg_b": stats_b["firstAvg"] or "-",
        "first_avg_delta": delta_ms(stats_a["firstAvgValue"], stats_b["firstAvgValue"]),
        "first_p90_a": stats_a["firstP90"] or "-",
        "first_p90_b": stats_b["firstP90"] or "-",
        "first_p90_delta": delta_ms(stats_a["firstP90Value"], stats_b["firstP90Value"]),
        "full_avg_a": stats_a["fullAvg"] or "-",
        "full_avg_b": stats_b["fullAvg"] or "-",
        "full_avg_delta": delta_ms(stats_a["fullAvgValue"], stats_b["fullAvgValue"]),
        "full_p90_a": stats_a["fullP90"] or "-",
        "full_p90_b": stats_b["fullP90"] or "-",
        "full_p90_delta": delta_ms(stats_a["fullP90Value"], stats_b["fullP90Value"]),
        "date_min_a": result_a.get("performance_summary", {}).get("date_min", ""),
        "date_max_a": result_a.get("performance_summary", {}).get("date_max", ""),
        "date_min_b": result_b.get("performance_summary", {}).get("date_min", ""),
        "date_max_b": result_b.get("performance_summary", {}).get("date_max", ""),
    }

    compare_modalities = compare_groups("modality", "-")
    compare_hangings = compare_groups("hanging", "Ohne Hanging")
    compare_users = compare_groups("user", "-")
    compare_chart_data = [
        {
            "label": row["label"],
            "countA": row["countA"],
            "countB": row["countB"],
            "firstAvgSecA": None if row.get("firstAvgMsA") is None else round(row["firstAvgMsA"] / 1000.0, 3),
            "firstAvgSecB": None if row.get("firstAvgMsB") is None else round(row["firstAvgMsB"] / 1000.0, 3),
            "fullAvgSecA": None if row.get("fullAvgMsA") is None else round(row["fullAvgMsA"] / 1000.0, 3),
            "fullAvgSecB": None if row.get("fullAvgMsB") is None else round(row["fullAvgMsB"] / 1000.0, 3),
            "firstP90SecA": None if row.get("firstP90MsA") is None else round(row["firstP90MsA"] / 1000.0, 3),
            "firstP90SecB": None if row.get("firstP90MsB") is None else round(row["firstP90MsB"] / 1000.0, 3),
            "fullP90SecA": None if row.get("fullP90MsA") is None else round(row["fullP90MsA"] / 1000.0, 3),
            "fullP90SecB": None if row.get("fullP90MsB") is None else round(row["fullP90MsB"] / 1000.0, 3),
            "slowA": row["slowA"],
            "slowB": row["slowB"],
        }
        for row in compare_modalities
    ]

    return {
        "xml_kind": "performancecsv_compare",
        "root": "performance.csv compare",
        "filename": f"{filename_a} ↔ {filename_b}",
        "filename_a": filename_a,
        "filename_b": filename_b,
        "performance_compare_label_a": label_a,
        "performance_compare_label_b": label_b,
        "missing_headers": sorted(set(result_a.get("missing_headers") or []) | set(result_b.get("missing_headers") or [])),
        "performance_percentile": PERFORMANCE_PERCENTILE,
        "performance_percentile_label": _performance_percentile_label(PERFORMANCE_PERCENTILE),
        "performance_percentile_hint": _performance_percentile_hint(PERFORMANCE_PERCENTILE),
        "performance_compare_summary": summary,
        "performance_compare_modalities": compare_modalities,
        "performance_compare_hangings": compare_hangings,
        "performance_compare_users": compare_users,
        "performance_compare_chart_data": compare_chart_data,
        "performance_a": result_a,
        "performance_b": result_b,
    }


def _split_mac_values(text):
    return [x.strip() for x in (text or "").split(",") if x.strip()]


def _analyze_license_file(root):
    licenses_node = root.find("licenses")
    issuer = root.find("./licenses/issuer")

    license_entries = licenses_node.findall("license") if licenses_node is not None else []
    server_identifications = [
        (x.text or "").strip()
        for x in root.findall("./licenses/serverIdentification")
        if (x.text or "").strip()
    ]

    id_counter = Counter((x.attrib.get("id", "") or "").strip().upper() for x in license_entries if x.attrib.get("id"))
    limitation_counter = Counter()
    parameter_counter = Counter()
    count_max_values = Counter()
    nodes_max_values = Counter()
    mac_values = []
    mac_description_index = defaultdict(set)
    mac_license_ids_index = defaultdict(set)
    mac_descriptions_by_license_id = defaultdict(lambda: defaultdict(set))

    for lic in license_entries:
        raw_desc_for_index = (lic.findtext("description") or "").strip().lower()
        macs_for_index = []
        for lim in lic.findall("limitation"):
            limit_on = (lim.attrib.get("limitOn", "") or "").strip()
            if limit_on:
                limitation_counter[limit_on] += 1
            for par in lim.findall("parameter"):
                pname = (par.attrib.get("name", "") or "").strip()
                pval = (par.text or "").strip()
                if pname:
                    parameter_counter[pname] += 1
                if limit_on == "MAC" and pname == "MAC" and pval:
                    lic_id = (lic.attrib.get("id", "") or "").strip().upper()
                    for mac in _split_mac_values(pval):
                        mac_values.append(mac)
                        macs_for_index.append(mac)
                        if lic_id:
                            mac_license_ids_index[mac].add(lic_id)
                            if raw_desc_for_index:
                                mac_descriptions_by_license_id[lic_id][mac].add(raw_desc_for_index)
                if limit_on == "COUNT" and pname == "MAX" and pval:
                    count_max_values[pval] += 1
                if limit_on == "NODES" and pname == "MAX" and pval:
                    nodes_max_values[pval] += 1

        if raw_desc_for_index:
            for m in macs_for_index:
                mac_description_index[m].add(raw_desc_for_index)

    focus_license_defs = [
        {"id": "DIAGNOST", "label": "Diagnost"},
        {"id": "XCHANGE_EXTENSIONS", "label": "XChange"},
        {"id": "XCHANGE", "label": "XChange (alt)"},
        {"id": "REVIEW", "label": "Review"},
        {"id": "DU_LIVE_VIEW_BASE", "label": "DeepUnity Live View"},
        {"id": "DU_LIVE_VIEW_AUDIO", "label": "DeepUnity Live View Audio"},
        {"id": "HECTEC", "label": "HecTec"},
        {"id": "IMPAXEE_MAMMO", "label": "ImpaxEE Mammo"},
        {"id": "TELERADIOLOGY", "label": "Teleradiologie"},
        {"id": "CLONE_VIEW", "label": "Clone View"},
        {"id": "DCMCDW_BURN", "label": "Dicom CDW-Burn"},
        {"id": "DICOM_PRINT", "label": "Dicom Print"},
        {"id": "HVR", "label": "HVR"},
        {"id": "WL_IMP", "label": "MWL Import"},
        {"id": "OFFLINE", "label": "Offline"},
        {"id": "LINK_ARCHIVE", "label": "Query Node"},
        {"id": "DICOM_SEND", "label": "Store Node"},
        {"id": "TEACHING_FILES", "label": "Teaching-Files"},
        {"id": "TOMTEC", "label": "TomTec"},
        {"id": "TOMTEC_ECHO", "label": "TomTec Cardiac"},
        {"id": "TOMTEC_MEASUREMENTS_ECHO", "label": "TomTec Cardiac Measurements"},
        {"id": "SCANNER", "label": "Vidar Scanner"},
        {"id": "VIDEO_PLAY", "label": "Video-Play"},
    ]
    known_focus_ids = {x["id"] for x in focus_license_defs}
    extra_ids = sorted([lid for lid in id_counter.keys() if lid and lid not in known_focus_ids])
    focus_license_defs.extend([{"id": lid, "label": f"{lid} (neu)"} for lid in extra_ids])
    focus_license_ids = [x["id"] for x in focus_license_defs]
    focus_license_label_by_id = {x["id"]: x["label"] for x in focus_license_defs}

    macs_by_focus_license = {k: defaultdict(lambda: {"counts": set(), "nodes": set(), "descriptions": set()}) for k in focus_license_ids}
    no_mac_by_focus_license = defaultdict(list)
    for lic in license_entries:
        lid = (lic.attrib.get("id", "") or "").strip().upper()
        if lid not in macs_by_focus_license:
            continue
        desc = _description_in_parentheses(lic.findtext("description"))
        count_values = {
            (par.text or "").strip()
            for lim in lic.findall("limitation")
            if (lim.attrib.get("limitOn", "") or "").strip().upper() == "COUNT"
            for par in lim.findall("parameter")
            if (par.attrib.get("name", "") or "").strip().upper() == "MAX" and (par.text or "").strip()
        }
        node_values = {
            (par.text or "").strip()
            for lim in lic.findall("limitation")
            if (lim.attrib.get("limitOn", "") or "").strip().upper() == "NODES"
            for par in lim.findall("parameter")
            if (par.attrib.get("name", "") or "").strip().upper() == "MAX" and (par.text or "").strip()
        }
        has_mac = False
        for lim in lic.findall("limitation"):
            if (lim.attrib.get("limitOn", "") or "").strip().upper() != "MAC":
                continue
            for par in lim.findall("parameter"):
                if (par.attrib.get("name", "") or "").strip().upper() != "MAC":
                    continue
                macs = _split_mac_values(par.text)
                for mac in macs:
                    has_mac = True
                    bucket = macs_by_focus_license[lid][mac]
                    if desc:
                        bucket["descriptions"].add(desc)
                    for c in count_values:
                        bucket["counts"].add(c)
                    for n in node_values:
                        bucket["nodes"].add(n)
        if not has_mac:
            no_mac_by_focus_license[lid].append({
                "description": desc,
                "counts": set(count_values),
                "nodes": set(node_values),
            })

    matrix_column_defs = [
        {"id": "DU_LIVE_VIEW_BASE", "label": "DeepUnity Live View"},
        {"id": "DU_LIVE_VIEW_AUDIO", "label": "DeepUnity Live View Audio"},
        {"id": "HECTEC", "label": "HecTec"},
        {"id": "IMPAXEE_MAMMO", "label": "ImpaxEE Mammo"},
        {"id": "TELERADIOLOGY", "label": "Teleradiologie"},
        {"id": "CLONE_VIEW", "label": "Clone View"},
        {"id": "DCMCDW_BURN", "label": "Dicom CDW-Burn (Deprecated)"},
        {"id": "DICOM_PRINT", "label": "Dicom Print"},
        {"id": "HVR", "label": "HVR"},
        {"id": "WL_IMP", "label": "MWL Import (Deprecated)"},
        {"id": "OFFLINE", "label": "Offline"},
        {"id": "LINK_ARCHIVE", "label": "Query Node"},
        {"id": "DICOM_SEND", "label": "Store Node"},
        {"id": "TEACHING_FILES", "label": "Teaching-Files (Deprecated)"},
        {"id": "TOMTEC", "label": "TomTec"},
        {"id": "TOMTEC_ECHO", "label": "TomTec Cardiac"},
        {"id": "TOMTEC_MEASUREMENTS_ECHO", "label": "TomTec Cardiac Measurements"},
        {"id": "SCANNER", "label": "Vidar Scanner"},
        {"id": "VIDEO_PLAY", "label": "Video-Play"},
    ]
    known_matrix_ids = {x["id"] for x in matrix_column_defs}
    matrix_column_defs.extend([
        {"id": lid, "label": f"{lid} (neu)"}
        for lid in extra_ids
        if lid not in known_matrix_ids
    ])
    server_based_ids = {"IMPAXFX"}
    matrix_columns = [
        x for x in matrix_column_defs
        if id_counter.get(x["id"], 0) > 0 and x["id"] not in server_based_ids
    ]
    boolean_only_license_ids = {
        "DCMCDW_BURN",
        "HVR",
        "OFFLINE",
        "TEACHING_FILES",
        "TOMTEC",
        "TOMTEC_ECHO",
        "TOMTEC_MEASUREMENTS_ECHO",
        "SCANNER",
    }
    def _capability_value(col_id, mac):
        lic_map = macs_by_focus_license.get(col_id, {})
        meta = lic_map.get(mac)
        if not meta:
            return "nein"
        if col_id in boolean_only_license_ids:
            return "ja"
        nodes = sorted(meta.get("nodes", []), key=lambda x: (len(x), x))
        if nodes:
            return ", ".join(nodes)
        counts = sorted(meta.get("counts", []), key=lambda x: (len(x), x))
        if counts:
            return ", ".join(counts)
        return "ja"

    def _has_any_license(mac, license_ids):
        for lid in license_ids:
            if mac in macs_by_focus_license.get(lid, {}):
                return True
        return False

    def _desc_blob_for(mac, license_ids):
        return " ".join(
            d
            for lid in license_ids
            for d in mac_descriptions_by_license_id.get(lid, {}).get(mac, set())
        ).lower()

    def _regular_r20_flag(mac):
        # Fachregel: DeepUnity R20 als regulärer Anteil
        return "deepunity r20" in _desc_blob_for(mac, {"IMPAXEE20", "IMPAXFX", "IMPAXEE_MAMMO", "IMPAXME"})

    def _regular_type_from_r20_desc(mac):
        blob = _desc_blob_for(mac, {"IMPAXEE20", "IMPAXFX", "IMPAXEE_MAMMO", "IMPAXME"})
        labels = []
        if "mammography" in blob or "mammo" in blob:
            labels.append("Mammography")
        if "review" in blob:
            labels.append("Review")
        if "pasta" in blob or " pa_" in blob or "(pa_" in blob:
            labels.append("Pasta")
        if "befund" in blob or "diagnost" in blob or " di_" in blob or "(di_" in blob:
            labels.append("Diagnost")
        # Fallback: wenn R20 vorhanden, aber kein Marker -> Diagnost als Standard
        if not labels and _regular_r20_flag(mac):
            labels.append("Diagnost")
        return labels

    legacy_type_ids = {"DIAGNOST", "REVIEW", "XCHANGE", "XCHANGE_EXTENSIONS"}

    def _diagnost_flag(mac):
        if _has_any_license(mac, {"DIAGNOST"}):
            return True
        return "Diagnost" in _regular_type_from_r20_desc(mac)

    def _review_flag(mac):
        if _has_any_license(mac, {"REVIEW"}):
            return True
        return "Review" in _regular_type_from_r20_desc(mac)

    def _review_advanced_flag(mac):
        if _has_any_license(mac, {"XCHANGE_EXTENSIONS"}):
            return True
        # in alten R20-Beschreibungen als Pasta geführt
        return "Pasta" in _regular_type_from_r20_desc(mac)

    def _deprecated_diagnostic_flag(mac):
        ids = mac_license_ids_index.get(mac, set())
        deprecated_markers = {"IMPAXEE20", "IMPAXFX", "IMPAXEE_MAMMO", "IMPAXME"}
        if ids & deprecated_markers:
            return True
        return False

    def _no_mac_count_for_license_ids(license_ids):
        total = 0
        for lid in license_ids:
            for entry in no_mac_by_focus_license.get(lid, []):
                numeric_values = []
                for value in entry.get("counts", set()):
                    value = (value or "").strip()
                    if value.isdigit():
                        numeric_values.append(int(value))
                if numeric_values:
                    total += sum(numeric_values)
                else:
                    total += 1
        return total

    def _deprecated_license_text(mac):
        deprecated_markers = {"IMPAXEE20", "IMPAXFX", "IMPAXEE_MAMMO", "IMPAXME"}
        if not _deprecated_diagnostic_flag(mac):
            return "-"
        ids_for_mac = mac_license_ids_index.get(mac, set())
        labels = []
        desc_blob = " ".join(
            d
            for lid in deprecated_markers
            for d in mac_descriptions_by_license_id.get(lid, {}).get(mac, set())
        ).lower()
        if "IMPAXEE_MAMMO" in ids_for_mac or "mammography" in desc_blob:
            labels.append("Mammography")
        if "diagnost" in desc_blob or "(di_" in desc_blob or " di_" in desc_blob:
            labels.append("Diagnost")
        if "review" in desc_blob or "(rv_" in desc_blob or " rv_" in desc_blob:
            labels.append("Review")
        if "(pa_" in desc_blob or " pa_" in desc_blob or "pasta" in desc_blob:
            labels.append("Pasta")
        if not labels:
            labels.extend(_regular_type_from_r20_desc(mac))
        return ", ".join(labels) if labels else "Unklar"

    focus_license_tables = [
        {
            "licenseId": lid,
            "licenseLabel": focus_license_label_by_id.get(lid, lid),
            "count": len(macs_by_focus_license[lid]),
            "rows": [
                {
                    "mac": mac,
                    "count": ", ".join(sorted(meta["counts"], key=lambda x: (len(x), x))) if meta["counts"] else "",
                    "description": " | ".join(sorted(meta["descriptions"])) if meta["descriptions"] else "",
                    "capabilities": {
                        col["id"]: _capability_value(col["id"], mac)
                        for col in matrix_columns
                    },
                }
                for mac, meta in sorted(macs_by_focus_license[lid].items(), key=lambda x: x[0])
            ],
        }
        for lid in focus_license_ids
    ]

    all_focus_macs = sorted({
        mac
        for lid in focus_license_ids
        for mac in macs_by_focus_license.get(lid, {}).keys()
    })

    license_matrix_groups = {}
    license_matrix_occurrences = defaultdict(int)
    for lic in license_entries:
        lid = (lic.attrib.get("id", "") or "").strip().upper()
        if lid not in focus_license_ids:
            continue
        desc = _description_in_parentheses(lic.findtext("description"))
        count_values = {
            (par.text or "").strip()
            for lim in lic.findall("limitation")
            if (lim.attrib.get("limitOn", "") or "").strip().upper() == "COUNT"
            for par in lim.findall("parameter")
            if (par.attrib.get("name", "") or "").strip().upper() == "MAX" and (par.text or "").strip()
        }
        node_values = {
            (par.text or "").strip()
            for lim in lic.findall("limitation")
            if (lim.attrib.get("limitOn", "") or "").strip().upper() == "NODES"
            for par in lim.findall("parameter")
            if (par.attrib.get("name", "") or "").strip().upper() == "MAX" and (par.text or "").strip()
        }
        for lim in lic.findall("limitation"):
            if (lim.attrib.get("limitOn", "") or "").strip().upper() != "MAC":
                continue
            for par in lim.findall("parameter"):
                if (par.attrib.get("name", "") or "").strip().upper() != "MAC":
                    continue
                for mac in _split_mac_values(par.text):
                    occurrence_key = (mac.lower(), desc.lower(), lid)
                    occurrence_index = license_matrix_occurrences[occurrence_key]
                    license_matrix_occurrences[occurrence_key] += 1
                    row_key = (mac, desc, occurrence_index)
                    group = license_matrix_groups.setdefault(row_key, {
                        "mac": mac,
                        "description": desc,
                        "index": occurrence_index,
                        "ids": set(),
                        "counts": set(),
                        "nodes": set(),
                        "by_id": defaultdict(lambda: {"counts": set(), "nodes": set()}),
                    })
                    group["ids"].add(lid)
                    group["counts"].update(count_values)
                    group["nodes"].update(node_values)
                    group["by_id"][lid]["counts"].update(count_values)
                    group["by_id"][lid]["nodes"].update(node_values)

    deprecated_r20_ids = {"IMPAXEE20", "IMPAXFX", "IMPAXEE_MAMMO", "IMPAXME"}

    def _row_regular_types(ids, description):
        blob = (description or "").lower()
        labels = []
        if "mammography" in blob or "mammo" in blob:
            labels.append("Mammography")
        if "review" in blob:
            labels.append("Review")
        if "pasta" in blob or " pa_" in blob or "(pa_" in blob:
            labels.append("Pasta")
        if "befund" in blob or "diagnost" in blob or " di_" in blob or "(di_" in blob:
            labels.append("Diagnost")
        if not labels and ids & deprecated_r20_ids and "deepunity r20" in blob:
            labels.append("Diagnost")
        return labels

    def _row_deprecated_text(ids, description):
        if not (ids & deprecated_r20_ids):
            return "-"
        blob = (description or "").lower()
        labels = []
        if "IMPAXEE_MAMMO" in ids or "mammography" in blob:
            labels.append("Mammography")
        if "diagnost" in blob or "(di_" in blob or " di_" in blob:
            labels.append("Diagnost")
        if "review" in blob or "(rv_" in blob or " rv_" in blob:
            labels.append("Review")
        if "(pa_" in blob or " pa_" in blob or "pasta" in blob:
            labels.append("Pasta")
        if not labels:
            labels.extend(_row_regular_types(ids, description))
        return ", ".join(labels) if labels else "Unklar"

    def _row_capability_value(group, col_id):
        if col_id not in group["ids"]:
            return "nein"
        if col_id in boolean_only_license_ids:
            return "ja"
        meta = group["by_id"].get(col_id, {})
        nodes = sorted(meta.get("nodes", []), key=lambda x: (len(x), x))
        if nodes:
            return ", ".join(nodes)
        counts = sorted(meta.get("counts", []), key=lambda x: (len(x), x))
        if counts:
            return ", ".join(counts)
        return "ja"

    license_matrix_rows = []
    license_matrix_duplicate_counter = Counter(
        (
            (group.get("mac") or "").strip().lower(),
            (group.get("description") or "").strip().lower(),
        )
        for group in license_matrix_groups.values()
        if (group.get("mac") or "").strip() and (group.get("description") or "").strip()
    )
    license_matrix_duplicate_keys = sorted([
        key for key, count in license_matrix_duplicate_counter.items() if count > 1
    ])
    license_matrix_duplicate_group_index = {
        key: idx for idx, key in enumerate(license_matrix_duplicate_keys)
    }
    for group in sorted(
        license_matrix_groups.values(),
        key=lambda x: ((x.get("mac") or "").lower(), (x.get("description") or "").lower(), x.get("index", 0)),
    ):
        ids = group["ids"]
        description = group.get("description", "")
        dup_key = (
            (group.get("mac") or "").strip().lower(),
            (description or "").strip().lower(),
        )
        is_duplicate = bool(dup_key[0] and dup_key[1] and license_matrix_duplicate_counter.get(dup_key, 0) > 1)
        row_regular_types = _row_regular_types(ids, description)
        diagnost = "DIAGNOST" in ids or "Diagnost" in row_regular_types
        review = "REVIEW" in ids or "Review" in row_regular_types
        review_advanced = "XCHANGE_EXTENSIONS" in ids or "Pasta" in row_regular_types
        all_counts = sorted(group["counts"], key=lambda x: (len(x), x))
        license_matrix_rows.append({
            "mac": group["mac"],
            "count": ", ".join(all_counts) if all_counts else "campus",
            "description": description,
            "isDuplicate": is_duplicate,
            "duplicateGroup": license_matrix_duplicate_group_index.get(dup_key) if is_duplicate else None,
            "diagnost": "ja" if diagnost else "nein",
            "review": "ja" if review else "nein",
            "review_advanced": "ja" if review_advanced else "nein",
            "deprecated": "ja" if ids & deprecated_r20_ids else "nein",
            "deprecated_license_text": _row_deprecated_text(ids, description),
            "license_text": ", ".join([
                lbl for lbl, ok in [
                    ("Diagnost", diagnost),
                    ("Review", review),
                    ("Review advanced", review_advanced),
                ] if ok
            ]) or "-",
            "capabilities": {
                col["id"]: _row_capability_value(group, col["id"])
                for col in matrix_columns
            },
        })

    # Campus-/Server-Lizenzen ohne MAC zusätzlich anzeigen, auch in gemischten Files.
    fallback_description_by_id = {
        "DIAGNOST": "Diagnost",
        "REVIEW": "Review",
        "XCHANGE_EXTENSIONS": "Review advanced",
        "IMPAXEE20": "Deprecated R20",
        "IMPAXFX": "Deprecated FX",
        "IMPAXEE_MAMMO": "Deprecated Mammography",
        "IMPAXME": "Deprecated Mobile Edition",
    }
    grouped = defaultdict(lambda: {"ids": set(), "counts": set(), "nodes": set()})
    for lid in focus_license_ids:
        for entry in no_mac_by_focus_license.get(lid, []):
            raw_desc = (entry.get("description", "") or "").strip()
            # Für reine Typ-Lizenzen ohne Klammerinhalt keine langen DeepUnity-* Texte anzeigen
            if lid in fallback_description_by_id and (not raw_desc or raw_desc.lower().startswith("deepunity ")):
                key = fallback_description_by_id[lid]
            else:
                key = raw_desc or focus_license_label_by_id.get(lid, lid) or "-"
            grouped[key]["ids"].add(lid)
            grouped[key]["counts"].update(entry.get("counts", set()))
            grouped[key]["nodes"].update(entry.get("nodes", set()))

    for desc, agg in sorted(grouped.items(), key=lambda x: x[0].lower()):
        ids = agg["ids"]
        counts = sorted(agg["counts"], key=lambda x: (len(x), x))
        nodes = sorted(agg["nodes"], key=lambda x: (len(x), x))

        def _cap_for(col_id):
            if col_id not in ids:
                return "nein"
            if col_id in boolean_only_license_ids:
                return "ja"
            if nodes:
                return ", ".join(nodes)
            if counts:
                return ", ".join(counts)
            return "ja"

        deprecated_flag = bool(ids & {"IMPAXEE20", "IMPAXFX", "IMPAXEE_MAMMO", "IMPAXME"})
        dep_type = "-"
        low = desc.lower()
        if deprecated_flag:
            if "mammo" in low or "mammography" in low:
                dep_type = "Mammography"
            elif "review" in low:
                dep_type = "Review"
            elif "pasta" in low or "pa_" in low:
                dep_type = "Pasta"
            elif "befund" in low or "diagnost" in low or "di_" in low:
                dep_type = "Diagnost"
            else:
                dep_type = "Unklar"

        license_matrix_rows.append({
            "mac": "-",
            "count": ", ".join(counts) if counts else "campus",
            "description": desc,
            "isDuplicate": False,
            "duplicateGroup": None,
            "diagnost": "ja" if "DIAGNOST" in ids else "nein",
            "review": "ja" if "REVIEW" in ids else "nein",
            "review_advanced": "ja" if "XCHANGE_EXTENSIONS" in ids else "nein",
            "deprecated": "ja" if deprecated_flag else "nein",
            "deprecated_license_text": dep_type,
            "license_text": ", ".join([x for x, ok in [
                ("Diagnost", "DIAGNOST" in ids),
                ("Review", "REVIEW" in ids),
                ("Review advanced", "XCHANGE_EXTENSIONS" in ids),
            ] if ok]) or "-",
            "capabilities": {col["id"]: _cap_for(col["id"]) for col in matrix_columns},
        })

    server_based_grouped = {"DataCare": 0, "Modalitäten": 0}
    for lic in license_entries:
        lid = ((lic.attrib.get("id", "") or "").strip().upper())
        text = (((lic.findtext("description") or "") + " " + lid)).lower()
        if "datacare" in text:
            server_based_grouped["DataCare"] += 1
        elif any(x in text for x in ["modalit", "modality", " modal"]):
            server_based_grouped["Modalitäten"] += 1

    server_based_licenses = [
        {"description": k, "count": v}
        for k, v in server_based_grouped.items()
        if v > 0
    ]
    license_overview = {
        "diagnost": (
            sum(1 for m in all_focus_macs if _diagnost_flag(m))
            + _no_mac_count_for_license_ids({"DIAGNOST"})
        ),
        "review": (
            sum(1 for m in all_focus_macs if _review_flag(m))
            + _no_mac_count_for_license_ids({"REVIEW"})
        ),
        "review_advanced": (
            sum(1 for m in all_focus_macs if _review_advanced_flag(m))
            + _no_mac_count_for_license_ids({"XCHANGE_EXTENSIONS"})
        ),
        "deprecated": (
            sum(1 for m in all_focus_macs if _deprecated_diagnostic_flag(m))
            + _no_mac_count_for_license_ids({"IMPAXEE20", "IMPAXFX", "IMPAXEE_MAMMO", "IMPAXME"})
        ),
    }

    return {
        "xml_kind": "licensefile",
        "root": root.tag,
        "issuer": {
            "name": issuer.attrib.get("name", "") if issuer is not None else "",
            "location": issuer.findtext("location", default="") if issuer is not None else "",
            "country": issuer.findtext("country", default="") if issuer is not None else "",
        },
        "signature": (root.findtext("signature") or "").strip(),
        "server_identifications": server_identifications,
        "license_stats": {
            "entries_total": len(license_entries),
            "types_total": len(id_counter),
            "mac_total": len(mac_values),
            "mac_unique": len(set(mac_values)),
        },
        "license_types": [
            {"id": lid, "count": count}
            for lid, count in sorted(id_counter.items(), key=lambda x: (-x[1], x[0]))
        ],
        "limitations_by_type": [
            {"limitOn": k, "count": v}
            for k, v in sorted(limitation_counter.items(), key=lambda x: (-x[1], x[0]))
        ],
        "parameters_by_name": [
            {"name": k, "count": v}
            for k, v in sorted(parameter_counter.items(), key=lambda x: (-x[1], x[0]))
        ],
        "count_max_values": [
            {"value": k, "count": v}
            for k, v in sorted(count_max_values.items(), key=lambda x: (-x[1], x[0]))
        ],
        "nodes_max_values": [
            {"value": k, "count": v}
            for k, v in sorted(nodes_max_values.items(), key=lambda x: (-x[1], x[0]))
        ],
        "focus_license_tables": focus_license_tables,
        "focus_matrix_columns": matrix_columns,
        "license_matrix_rows": license_matrix_rows,
        "license_overview": license_overview,
        "server_based_licenses": server_based_licenses,
    }


def analyze_xml(file_storage):
    data = file_storage.read()
    root = ET.fromstring(data)

    xml_kind = _detect_xml_kind(root)
    if xml_kind == "licensefile":
        return _analyze_license_file(root)

    roles = []

    def walk(node, depth=1, path=""):
        for r in node.findall("role"):
            name = r.attrib.get("name", "")
            rp = (path + "/" + name).strip("/")
            cfg = r.find("configuration")
            items = cfg.findall("item") if cfg is not None else []
            roles.append({"path": rp, "depth": depth, "items": len(items), "item_nodes": items})
            walk(r, depth + 1, rp)

    walk(root)

    all_items = []
    for cfg in root.iter("configuration"):
        all_items.extend(cfg.findall("item"))

    type_counts = Counter(i.attrib.get("type", "") for i in all_items)

    # DICOM send endpoints grouped by ID. Keep the role(s) where every endpoint is configured,
    # because the global item scan below otherwise loses the role context.
    endpoints = defaultdict(dict)
    endpoint_roles = defaultdict(set)
    role_item_ids = set()
    prefix = "listtext.service.dicomsend#"

    def _dicom_send_item_parts(item):
        name = _normalize_item_name(item.attrib.get("name", ""))
        if not name.startswith(prefix):
            return None, None
        rest = name[len(prefix):]
        if "." not in rest:
            return None, None
        return rest.split(".", 1)

    for r in roles:
        for i in r.get("item_nodes", []):
            role_item_ids.add(id(i))
            eid, key = _dicom_send_item_parts(i)
            if not eid:
                continue
            endpoint_roles[eid].add(r.get("path", "") or "(ohne Rolle)")

    for i in all_items:
        eid, key = _dicom_send_item_parts(i)
        if not eid:
            continue
        endpoints[eid][key] = (i.text or "").strip()
        if id(i) not in role_item_ids:
            endpoint_roles[eid].add("(global / keine Rolle)")

    def _join_display_values(values, sep=" | "):
        out = []
        seen = set()
        for v in values:
            t = (v or "").strip()
            if not t:
                continue
            low = t.lower()
            if low in seen:
                continue
            seen.add(low)
            out.append(t)
        return sep.join(out)

    dicom_endpoints = []
    for eid, vals in sorted(endpoints.items(), key=lambda x: x[0]):
        roles_for_endpoint = sorted(endpoint_roles.get(eid, []), key=lambda x: x.lower())
        dicom_endpoints.append({
            "id": eid,
            "role": _join_display_values(roles_for_endpoint),
            "roles": roles_for_endpoint,
            "description": vals.get("description", ""),
            "host": vals.get("host", ""),
            "port": vals.get("port", ""),
            "callingAET": vals.get("callingAET", ""),
            "calledAET": vals.get("calledAET", ""),
            "secure": vals.get("secure", ""),
            "use_cmove": vals.get("use_cmove", ""),
            "close_on_each": vals.get("close_on_each", ""),
            "disableDirectImport": vals.get("disableDirectImport", ""),
        })

    def _endpoint_dup_norm(v: str) -> str:
        return (v or "").strip().lower()

    endpoint_dup_counter = Counter(
        (
            _endpoint_dup_norm(e.get("host")),
            _endpoint_dup_norm(e.get("calledAET")),
        )
        for e in dicom_endpoints
        if _endpoint_dup_norm(e.get("host")) and _endpoint_dup_norm(e.get("calledAET"))
    )
    endpoint_duplicate_keys = sorted([k for k, c in endpoint_dup_counter.items() if c > 1])
    endpoint_duplicate_group_index = {k: idx for idx, k in enumerate(endpoint_duplicate_keys)}

    for e in dicom_endpoints:
        dup_key = (
            _endpoint_dup_norm(e.get("host")),
            _endpoint_dup_norm(e.get("calledAET")),
        )
        is_dup = bool(dup_key[0] and dup_key[1] and endpoint_dup_counter.get(dup_key, 0) > 1)
        e["isDuplicate"] = is_dup
        e["duplicateGroup"] = endpoint_duplicate_group_index.get(dup_key) if is_dup else None

    # Keep endpoints with the same duplicate color/group next to each other.
    # Duplicate groups are shown first, followed by the remaining endpoints.
    dicom_endpoints.sort(key=lambda e: (
        0 if e.get("isDuplicate") else 1,
        e.get("duplicateGroup") if e.get("duplicateGroup") is not None else 999999,
        _endpoint_dup_norm(e.get("host")),
        _endpoint_dup_norm(e.get("calledAET")),
        (e.get("description") or "").strip().lower(),
        (e.get("role") or "").strip().lower(),
        str(e.get("id") or ""),
    ))

    # Archives / data nodes grouped by ID. Keep the role(s) where every archive is
    # configured, just like for the DICOM send endpoints above.
    data_nodes = defaultdict(dict)
    archive_roles = defaultdict(set)
    role_data_item_ids = set()
    data_prefix = "listtext.datanode.datanodes#"

    def _data_node_item_parts(item):
        name = _normalize_item_name(item.attrib.get("name", ""))
        if not name.startswith(data_prefix):
            return None, None
        rest = name[len(data_prefix):]
        if "." not in rest:
            return None, None
        return rest.split(".", 1)

    for r in roles:
        for i in r.get("item_nodes", []):
            nid, key = _data_node_item_parts(i)
            if not nid:
                continue
            role_data_item_ids.add(id(i))
            archive_roles[nid].add(r.get("path", "") or "(ohne Rolle)")

    for i in all_items:
        nid, key = _data_node_item_parts(i)
        if not nid:
            continue
        data_nodes[nid][key] = (i.text or "").strip()
        if id(i) not in role_data_item_ids:
            archive_roles[nid].add("(global / keine Rolle)")

    archives = []
    for nid, vals in sorted(data_nodes.items(), key=lambda x: x[0]):
        archive_name = vals.get("name", "")
        if not archive_name:
            continue
        roles_for_archive = sorted(archive_roles.get(nid, []), key=lambda x: x.lower())
        ref_count = sum(1 for i in all_items if archive_name and archive_name in ((i.text or "").strip()))
        archives.append({
            "id": nid,
            "name": archive_name,
            "description": archive_name,
            "role": _join_display_values(roles_for_archive),
            "roles": roles_for_archive,
            "type": vals.get("type", ""),
            "host": vals.get("properties.host", ""),
            "port": vals.get("properties.port", ""),
            "secureConnection": vals.get("properties.secureConnection", ""),
            "calledAET": vals.get("properties.calledAET", ""),
            "callingAET": vals.get("properties.callingAET", ""),
            "callingAET2": vals.get("properties.callingAET2", ""),
            "automaticStudyMonitoring": vals.get("properties.automaticStudyMonitoring", ""),
            "useCGet": vals.get("properties.useCGet", ""),
            "references": ref_count,
        })

    def _archive_dup_norm(v: str) -> str:
        return (v or "").strip().lower()

    archive_dup_counter = Counter(
        (
            _archive_dup_norm(a.get("host")),
            _archive_dup_norm(a.get("calledAET")),
        )
        for a in archives
        if _archive_dup_norm(a.get("host")) and _archive_dup_norm(a.get("calledAET"))
    )
    archive_duplicate_keys = sorted([k for k, c in archive_dup_counter.items() if c > 1])
    archive_duplicate_group_index = {k: idx for idx, k in enumerate(archive_duplicate_keys)}

    for a in archives:
        dup_key = (
            _archive_dup_norm(a.get("host")),
            _archive_dup_norm(a.get("calledAET")),
        )
        is_dup = bool(dup_key[0] and dup_key[1] and archive_dup_counter.get(dup_key, 0) > 1)
        a["isDuplicate"] = is_dup
        a["duplicateGroup"] = archive_duplicate_group_index.get(dup_key) if is_dup else None

    # Keep duplicate archive groups next to each other, using the same approach as
    # Speichern and DICOM-Printers.
    archives.sort(key=lambda a: (
        0 if a.get("isDuplicate") else 1,
        a.get("duplicateGroup") if a.get("duplicateGroup") is not None else 999999,
        _archive_dup_norm(a.get("host")),
        _archive_dup_norm(a.get("calledAET")),
        (a.get("name") or "").strip().lower(),
        (a.get("role") or "").strip().lower(),
        str(a.get("id") or ""),
    ))

    enterprise_roles = [
        {"path": r["path"], "depth": r["depth"], "items": r["items"]}
        for r in roles if r["path"] == "Enterprise" or r["path"].startswith("Enterprise/")
    ]
    workstation_roles = [
        {"path": r["path"], "depth": r["depth"], "items": r["items"]}
        for r in roles if r["path"].startswith("Workstations/")
    ]

    role_by_path = {r["path"]: r for r in roles}

    def _resolve_permission(role_path, key):
        current = role_path
        while current:
            role = role_by_path.get(current)
            if role:
                for item in role.get("item_nodes", []):
                    item_name = _normalize_item_name(item.attrib.get("name", ""))
                    if item_name == key:
                        value = (item.text or "").strip()
                        if current == role_path:
                            return value, "direct", current
                        return value, "inherited", current
            current = current.rsplit("/", 1)[0] if "/" in current else ""
        return "", "missing", ""

    permission_labels = _load_permission_label_map()
    all_permission_keys = sorted({
        _normalize_item_name(i.attrib.get("name", ""))
        for i in all_items
        if _normalize_item_name(i.attrib.get("name", "")).startswith("listtext.permissions")
    })

    enterprise_permission_rows = []
    scoped_permission_keys = [k for k in all_permission_keys if k in PERMISSION_PATH_OVERRIDES]

    for er in enterprise_roles:
        role_path = er["path"]
        for key in scoped_permission_keys:
            value, source_type, source_role = _resolve_permission(role_path, key)
            if source_type == "missing":
                value = "false"
                source_type = "direct"
                source_role = role_path
                state = "verboten"
            elif source_type == "inherited":
                state = "vererbt"
            else:
                state = "erlaubt" if value.lower() == "true" else "verboten"

            display_path = _permission_display_path(key, permission_labels)
            de_segments = [x.strip() for x in display_path.split("/") if x.strip()]

            enterprise_permission_rows.append({
                "role": role_path,
                "label": de_segments[-1] if de_segments else key,
                "key": key,
                "displayPath": display_path,
                "value": value,
                "sourceType": source_type,
                "sourceRole": source_role,
                "state": state,
            })

    enterprise_permission_roles = sorted({x["role"] for x in enterprise_permission_rows})

    workstation_item_details = []
    workstation_devices = []
    stale_cutoff = datetime.now() - timedelta(days=90)

    ws_setting_keys = {
        "otherSettings": "listtext.workstation.webdeployer.extended.otherSettings",
        "loggingLevel": "logging.level",
        "installContainer": "listtext.workstation.webdeployer.basic.Install.Container",
        "applicationType": "listtext.workstation.webdeployer.basic.Starter.ApplicationType",
        "teleradiology": "hardware.TELERADIOLOGY",
    }

    for r in roles:
        if not r["path"].startswith("Workstations/"):
            continue

        role_settings = {k: "" for k in ws_setting_keys}
        for item in r["item_nodes"]:
            item_name = _normalize_item_name(item.attrib.get("name", ""))
            item_value = (item.text or "").strip()
            for key, full_name in ws_setting_keys.items():
                if item_name == full_name:
                    role_settings[key] = item_value

        for item in r["item_nodes"]:
            name = _normalize_item_name(item.attrib.get("name", ""))
            value = (item.text or "").strip()
            workstation_item_details.append({
                "path": r["path"],
                "name": name,
                "type": item.attrib.get("type", ""),
                "value": value,
            })

            if name == "listtext.hardware.information" and value:
                try:
                    hw = ET.fromstring(value)
                    last_updated = hw.attrib.get("lastUpdated", "")
                    last_updated_dt = _parse_last_updated(last_updated)
                    last_updated_date_key = last_updated_dt.strftime("%Y-%m-%d") if last_updated_dt else ""
                    is_stale = bool(last_updated_dt and last_updated_dt < stale_cutoff)

                    graphics_adapters = ", ".join(
                        x.attrib.get("name", "") for x in hw.findall("graphicsAdapter") if x.attrib.get("name", "")
                    )
                    monitors = ", ".join(
                        x.attrib.get("name", "") for x in hw.findall("monitor") if x.attrib.get("name", "")
                    )

                    hardware_xml_lines = _flatten_xml_values(hw)
                    hardware_xml_tooltip = "\n".join(hardware_xml_lines)

                    workstation_devices.append({
                        "path": r["path"],
                        "computerName": hw.attrib.get("computerName", ""),
                        "configuredAET": hw.attrib.get("configuredAET", ""),
                        "ip": hw.attrib.get("ip", ""),
                        "mac": hw.attrib.get("mac", ""),
                        "os": hw.attrib.get("os", ""),
                        "cpu": hw.attrib.get("cpu", ""),
                        "graphicsAdapter": graphics_adapters,
                        "monitor": monitors,
                        "ramSize": hw.attrib.get("ramSize", ""),
                        "serialNumber": hw.attrib.get("serialNumber", ""),
                        "softwareVersion": hw.attrib.get("softwareVersion", ""),
                        "availableLicenses": hw.attrib.get("availableLicenses", ""),
                        "lastUpdated": last_updated,
                        "lastUpdatedDateKey": last_updated_date_key,
                        "isStale": is_stale,
                        "otherSettings": role_settings["otherSettings"],
                        "loggingLevel": role_settings["loggingLevel"],
                        "installContainer": role_settings["installContainer"],
                        "applicationType": role_settings["applicationType"],
                        "teleradiology": role_settings["teleradiology"],
                        "hardwareXmlTooltip": hardware_xml_tooltip,
                    })
                except Exception:
                    pass

    workstation_item_details.sort(key=lambda x: (x["path"], x["name"]))
    workstation_devices.sort(key=lambda x: x["path"])

    raw_workstation_devices = workstation_devices

    def _split_csv_vals(v):
        return [x.strip() for x in (v or "").split(",") if x and x.strip()]

    def _join_unique(values, sep=", "):
        out = []
        seen = set()
        for v in values:
            t = (v or "").strip()
            if not t:
                continue
            low = t.lower()
            if low in seen:
                continue
            seen.add(low)
            out.append(t)
        return sep.join(out)

    grouped_ws = defaultdict(list)
    for d in raw_workstation_devices:
        key = (
            (d.get("computerName", "") or "").strip().lower(),
            (d.get("configuredAET", "") or "").strip().lower(),
        )
        grouped_ws[key].append(d)

    merged_workstation_devices = []
    for _, rows in grouped_ws.items():
        rows_sorted = sorted(rows, key=lambda x: x.get("path", ""))
        merged = {
            "path": _join_unique([r.get("path", "") for r in rows_sorted], sep=" | "),
            "computerName": _join_unique([r.get("computerName", "") for r in rows_sorted], sep=" | "),
            "configuredAET": _join_unique([r.get("configuredAET", "") for r in rows_sorted], sep=" | "),
            "ip": _join_unique([r.get("ip", "") for r in rows_sorted]),
            "mac": _join_unique([r.get("mac", "") for r in rows_sorted]),
            "os": _join_unique([r.get("os", "") for r in rows_sorted], sep=" | "),
            "cpu": _join_unique([r.get("cpu", "") for r in rows_sorted], sep=" | "),
            "graphicsAdapter": _join_unique([r.get("graphicsAdapter", "") for r in rows_sorted], sep=" | "),
            "monitor": _join_unique([r.get("monitor", "") for r in rows_sorted], sep=" | "),
            "ramSize": _join_unique([r.get("ramSize", "") for r in rows_sorted]),
            "serialNumber": _join_unique([r.get("serialNumber", "") for r in rows_sorted]),
            "softwareVersion": _join_unique([r.get("softwareVersion", "") for r in rows_sorted]),
            "availableLicenses": _join_unique(
                sorted({lic for r in rows_sorted for lic in _split_csv_vals(r.get("availableLicenses", ""))})
            ),
            "lastUpdated": _join_unique([r.get("lastUpdated", "") for r in rows_sorted], sep=" | "),
            "lastUpdatedDateKeys": _join_unique([r.get("lastUpdatedDateKey", "") for r in rows_sorted], sep=" "),
            "isStale": any(bool(r.get("isStale")) for r in rows_sorted),
            "otherSettings": _join_unique([r.get("otherSettings", "") for r in rows_sorted], sep=" | "),
            "loggingLevel": _join_unique([r.get("loggingLevel", "") for r in rows_sorted], sep=" | "),
            "installContainer": _join_unique([r.get("installContainer", "") for r in rows_sorted], sep=" | "),
            "applicationType": _join_unique([r.get("applicationType", "") for r in rows_sorted], sep=" | "),
            "teleradiology": _join_unique([r.get("teleradiology", "") for r in rows_sorted], sep=" | "),
            "hardwareXmlTooltip": "\n\n---\n\n".join([r.get("hardwareXmlTooltip", "") for r in rows_sorted if r.get("hardwareXmlTooltip", "")]),
            "rawCount": len(rows_sorted),
            "hasDuplicate": len(rows_sorted) > 1,
        }
        merged_workstation_devices.append(merged)

    merged_workstation_devices.sort(key=lambda x: x.get("path", ""))
    workstation_devices = merged_workstation_devices

    workstation_license_columns = sorted({
        lic
        for d in workstation_devices
        for lic in _split_csv_vals(d.get("availableLicenses", ""))
        if lic
    })

    workstation_license_rows = []
    for d in workstation_devices:
        present = set(_split_csv_vals(d.get("availableLicenses", "")))
        workstation_license_rows.append({
            "computerName": d.get("computerName", ""),
            "configuredAET": d.get("configuredAET", ""),
            "mac": d.get("mac", ""),
            "rawCount": d.get("rawCount", 1),
            "hasDuplicate": bool(d.get("hasDuplicate", False)),
            "licenseFlags": {
                lic: ("ja" if lic in present else "nein")
                for lic in workstation_license_columns
            },
        })

    script_map = defaultdict(dict)
    csv_hits = []
    script_prefix = "listtext.scripts#"

    for i in all_items:
        name = _normalize_item_name(i.attrib.get("name", ""))
        value = (i.text or "").strip()

        if name.startswith(script_prefix):
            rest = name[len(script_prefix):]
            if "." in rest:
                sid, key = rest.split(".", 1)
                script_map[sid][key] = value

        if "csv" in name.lower() or "csv" in value.lower():
            csv_hits.append({
                "name": name,
                "type": i.attrib.get("type", ""),
                "value": value[:300],
            })

    scripts = []
    for sid, vals in sorted(script_map.items(), key=lambda x: x[0]):
        code = vals.get("code", "")
        svg_icon_code = vals.get("svgIcon.code", "")
        svg_preview_data_uri = ""
        svg_icon_decoded_len = 0

        if svg_icon_code:
            try:
                decoded_bytes = gzip.decompress(base64.b64decode(svg_icon_code))
                decoded_text = decoded_bytes.decode("utf-8", errors="replace")
                svg_icon_decoded_len = len(decoded_text)
                if "<svg" in decoded_text:
                    svg_preview_data_uri = "data:image/svg+xml;base64," + base64.b64encode(decoded_bytes).decode("ascii")
                svg_text = decoded_text
            except Exception:
                svg_text = ""
                pass
        else:
            svg_text = ""

        icon_name = vals.get("svgIcon.name", "")

        scripts.append({
            "id": sid,
            "name": vals.get("name", ""),
            "showInLTA": vals.get("showInLTA", ""),
            "hasCode": bool(code),
            "fullCode": code,
            "icon": icon_name,
            "svgDownloadName": _safe_svg_filename(icon_name, f"script_{sid}"),
            "svgIconCodeLength": len(svg_icon_code),
            "svgIconDecodedLength": svg_icon_decoded_len,
            "svgPreviewDataUri": svg_preview_data_uri,
            "svgText": svg_text,
        })

    descriptor_meta = {}
    descriptor_raw_xml = {}
    descriptor_prefix = "impaxee.jvision.descriptors#"
    for i in all_items:
        name = _normalize_item_name(i.attrib.get("name", ""))
        if not name.startswith(descriptor_prefix) or not name.endswith(".content"):
            continue

        raw_xml = (i.text or "").strip()
        if not raw_xml:
            continue

        try:
            droot = ET.fromstring(raw_xml)
            did = (droot.attrib.get("id", "") or "").strip()
            if not did:
                continue

            tags = set()
            for node in droot.iter():
                tag_attr = (node.attrib.get("tag", "") or "").strip()
                if tag_attr:
                    tags.add(tag_attr)
                criterion_tag = (node.attrib.get("criterion_tag", "") or "").strip()
                if criterion_tag:
                    tags.add(criterion_tag)

            operators = []
            for node in droot.iter():
                op = (node.attrib.get("operator", "") or "").strip()
                if op and op not in operators:
                    operators.append(op)

            conditions = []
            for node in droot.iter():
                op = (node.attrib.get("operator", "") or "").strip()
                tag_attr = (node.attrib.get("tag", "") or "").strip() or (node.attrib.get("criterion_tag", "") or "").strip()
                value_attr = (node.attrib.get("value", "") or "").strip()
                if op or tag_attr or value_attr:
                    conditions.append({
                        "operator": op,
                        "tag": tag_attr,
                        "value": value_attr,
                    })

            descriptor_meta[did] = {
                "name": droot.attrib.get("name", ""),
                "tags": sorted(tags),
                "operators": operators,
                "conditions": conditions,
            }
            descriptor_raw_xml[did] = raw_xml
        except Exception:
            continue

    hanging_protocols = []
    hanging_condition_set_rows = []
    _hanging_condition_set_seen = set()
    hanging_prefix = "impaxee.jvision.HANGMAN.hangingProtocols#"
    for i in all_items:
        name = _normalize_item_name(i.attrib.get("name", ""))
        if not name.startswith(hanging_prefix) or not name.endswith(".content"):
            continue

        raw_xml = (i.text or "").strip()
        if not raw_xml:
            continue

        protocol_id = name[len(hanging_prefix):].split(".", 1)[0]
        try:
            hp = ET.fromstring(raw_xml)
            hp_name = (hp.attrib.get("name", "") or "").strip()
            hp_export_name = hp_name if hp_name.startswith("EXT_") else f"EXT_{hp_name}"
            export_raw_xml = re.sub(
                r"(<hangingProtocol\b[^>]*\bname=')([^']*)(')",
                lambda m: m.group(1) + hp_export_name + m.group(3),
                raw_xml,
                count=1,
            )

            display_descriptor_ids = sorted({
                (ref.attrib.get("id", "") or "").strip()
                for ref in hp.findall('.//hangingDefinition//ReferencedDescriptor')
                if (ref.attrib.get("id", "") or "").strip()
            })

            display_descriptor_names = []
            display_dicom_tags = []
            for did in display_descriptor_ids:
                meta = descriptor_meta.get(did)
                if not meta:
                    continue
                dname = (meta.get("name", "") or "").strip()
                if dname and dname not in display_descriptor_names:
                    display_descriptor_names.append(dname)
                for t in meta.get("tags", []):
                    if t not in display_dicom_tags:
                        display_dicom_tags.append(t)

            conditional_descriptor_ids = sorted({
                (ref.attrib.get("id", "") or "").strip()
                for ref in hp.findall('.//snapshot//conditionalHanging//displaySetCondition//ReferencedDescriptor')
                if (ref.attrib.get("id", "") or "").strip()
            })
            for did in conditional_descriptor_ids:
                meta = descriptor_meta.get(did) or {}
                dname = (meta.get("name", "") or "").strip() or did
                for cond in (meta.get("conditions") or []):
                    op = (cond.get("operator", "") or "").strip()
                    tag = (cond.get("tag", "") or "").strip()
                    value = (cond.get("value", "") or "").strip()
                    key = (dname, op, tag, value)
                    if key in _hanging_condition_set_seen:
                        continue
                    _hanging_condition_set_seen.add(key)
                    tag_label = _dicom_tag_to_hex_label(tag)
                    tag_desc, tag_number = _split_dicom_label(tag_label)
                    descriptor_xml = descriptor_raw_xml.get(did, "")
                    agd_data_uri = _build_agd_data_uri_from_descriptors([descriptor_xml]) if descriptor_xml else ""
                    hanging_condition_set_rows.append({
                        "descriptorId": did,
                        "name": dname,
                        "condition": op,
                        "dicomTagDescription": tag_desc,
                        "dicomTagNumber": tag_number,
                        "value": value,
                        "agdDataUri": agd_data_uri,
                        "agdDownloadName": _safe_download_filename(dname, f"condition_{did}", ".AGD"),
                    })

            all_descriptor_ids = sorted({
                (ref.attrib.get("id", "") or "").strip()
                for ref in hp.findall('.//ReferencedDescriptor')
                if (ref.attrib.get("id", "") or "").strip()
            })
            descriptor_blocks = [descriptor_raw_xml[did] for did in all_descriptor_ids if did in descriptor_raw_xml]

            snapshot_views = []
            for snap in hp.findall('.//snapshot'):
                snap_name = (snap.attrib.get('name', '') or '').strip() or '(ohne Namen)'
                screens = []

                # 1) Aus conditionalHangingScreen lesen
                snapshot_screen_layouts = []
                for sl in snap.findall('.//screenLayout'):
                    hv = (sl.attrib.get('horizontalXVertical', '') or '').strip()
                    if hv:
                        snapshot_screen_layouts.append(hv)

                for cs in snap.findall('.//conditionalHangingScreen'):
                    screen_name = (
                        (cs.attrib.get('name', '') or '').strip()
                        or (cs.attrib.get('screen', '') or '').strip()
                        or f"Screen {len(screens) + 1}"
                    )

                    layouts = []
                    for sl in cs.findall('.//screenLayout'):
                        hv = (sl.attrib.get('horizontalXVertical', '') or '').strip()
                        if hv:
                            layouts.append(hv)
                    for dl in cs.findall('.//displayLayout'):
                        hv = (dl.attrib.get('horizontalXVertical', '') or '').strip()
                        if hv:
                            layouts.append(hv)
                    if not layouts:
                        for l in cs.findall('.//layout'):
                            t = (l.attrib.get('type', '') or '').strip()
                            if t:
                                layouts.append(t)

                    # Häufig liegt das eigentliche Screen-Muster am snapshot-level
                    if not layouts and snapshot_screen_layouts:
                        idx = len(screens)
                        if idx < len(snapshot_screen_layouts):
                            layouts = [snapshot_screen_layouts[idx]]
                        else:
                            layouts = [snapshot_screen_layouts[-1]]

                    condition_names = []
                    condition_rules = []
                    for idx_dsc, dsc in enumerate(cs.findall('.//displaySetCondition'), 1):
                        n = (dsc.attrib.get('name', '') or '').strip()
                        if not n:
                            n = (dsc.findtext('name', '') or '').strip()
                        if not n:
                            n = f"Regel {idx_dsc}"
                        condition_names.append(n)
                        rule_operators = []
                        rule_condition_names = []
                        for ref in dsc.findall('.//ReferencedDescriptor'):
                            rid = (ref.attrib.get('id', '') or '').strip()
                            if not rid:
                                continue
                            rmeta = descriptor_meta.get(rid, {})
                            rname = (rmeta.get('name', '') or '').strip()
                            if rname and rname not in rule_condition_names:
                                rule_condition_names.append(rname)
                            for op in descriptor_meta.get(rid, {}).get('operators', []):
                                if op and op not in rule_operators:
                                    rule_operators.append(op)
                        condition_rules.append({
                            'label': f"Regel {idx_dsc}",
                            'name': n,
                            'conditionNames': rule_condition_names,
                            'operators': rule_operators,
                        })

                    if not condition_names:
                        for idx_dsc, dsc in enumerate(snap.findall('.//displaySetCondition'), 1):
                            n = (dsc.attrib.get('name', '') or '').strip()
                            if not n:
                                n = (dsc.findtext('name', '') or '').strip()
                            if not n:
                                n = f"Regel {idx_dsc}"
                            condition_names.append(n)
                            rule_operators = []
                            rule_condition_names = []
                            for ref in dsc.findall('.//ReferencedDescriptor'):
                                rid = (ref.attrib.get('id', '') or '').strip()
                                if not rid:
                                    continue
                                rmeta = descriptor_meta.get(rid, {})
                                rname = (rmeta.get('name', '') or '').strip()
                                if rname and rname not in rule_condition_names:
                                    rule_condition_names.append(rname)
                                for op in descriptor_meta.get(rid, {}).get('operators', []):
                                    if op and op not in rule_operators:
                                        rule_operators.append(op)
                            condition_rules.append({
                                'label': f"Regel {idx_dsc}",
                                'name': n,
                                'conditionNames': rule_condition_names,
                                'operators': rule_operators,
                            })

                    condition_name = ', '.join(dict.fromkeys(condition_names))

                    parsed_layouts = []
                    for lay in layouts:
                        text = (lay or '').strip()
                        m = re.match(r'^(\d+)x(\d+)$', text.lower())
                        if m:
                            parsed_layouts.append({
                                'label': text,
                                'cols': int(m.group(1)),
                                'rows': int(m.group(2)),
                                'conditionName': condition_name,
                                'conditionRules': condition_rules,
                            })
                        else:
                            parsed_layouts.append({
                                'label': text,
                                'cols': None,
                                'rows': None,
                                'conditionName': condition_name,
                                'conditionRules': condition_rules,
                            })

                    screens.append({
                        'name': screen_name,
                        'layouts': layouts,
                        'layoutTables': parsed_layouts,
                    })

                # 2) Fallback: layout direkt unter snapshot
                if not screens:
                    fallback_layouts = []
                    fallback_layouts.extend(snapshot_screen_layouts)
                    for dl in snap.findall('.//displayLayout'):
                        hv = (dl.attrib.get('horizontalXVertical', '') or '').strip()
                        if hv:
                            fallback_layouts.append(hv)
                    if fallback_layouts:
                        parsed_layouts = []
                        for lay in fallback_layouts:
                            text = (lay or '').strip()
                            m = re.match(r'^(\d+)x(\d+)$', text.lower())
                            if m:
                                parsed_layouts.append({
                                    'label': text,
                                    'cols': int(m.group(1)),
                                    'rows': int(m.group(2)),
                                    'conditionName': '',
                                    'conditionRules': [],
                                })
                            else:
                                parsed_layouts.append({
                                    'label': text,
                                    'cols': None,
                                    'rows': None,
                                    'conditionName': '',
                                    'conditionRules': [],
                                })
                        screens.append({
                            'name': 'Screen 1',
                            'layouts': fallback_layouts,
                            'layoutTables': parsed_layouts,
                        })

                snapshot_views.append({
                    'name': snap_name,
                    'screens': screens,
                    'hasConditionalHanging': snap.find('.//conditionalHanging') is not None,
                    'hasOrderHanging': snap.find('.//orderHanging') is not None,
                })

            agh_xml_payload = (
                "<hangingImportExport>\n"
                + ("\n".join(descriptor_blocks) + "\n" if descriptor_blocks else "")
                + export_raw_xml
                + "\n</hangingImportExport>"
            )
            agh_xml_bytes = agh_xml_payload.encode("utf-8")
            agh_signature = _calc_agh_signature(agh_xml_bytes)
            agh_binary = _build_agh_binary(agh_xml_payload, token=agh_signature)
            agh_data_uri = "data:application/octet-stream;base64," + base64.b64encode(
                agh_binary
            ).decode("ascii")

            hanging_protocols.append({
                "id": protocol_id,
                "name": hp.attrib.get("name", ""),
                "active": hp.attrib.get("active", ""),
                "numberOfPriors": hp.attrib.get("numberOfPriors", ""),
                "priority": hp.attrib.get("priority", ""),
                "usePriors": hp.attrib.get("usePriors", ""),
                "createdBy": hp.attrib.get("createdBy", ""),
                "lastModifiedBy": hp.attrib.get("lastModifiedBy", ""),
                "displayDescriptorNames": ", ".join(display_descriptor_names),
                "displayDicomTags": ", ".join(display_dicom_tags),
                "snapshotViews": snapshot_views,
                "aghDataUri": agh_data_uri,
                "aghDownloadName": _safe_download_filename(
                    hp_export_name,
                    f"EXT_hanging_{protocol_id}",
                    ".AGH"
                ),
            })
        except Exception:
            hanging_protocols.append({
                "id": protocol_id,
                "name": "(nicht lesbar)",
                "active": "",
                "numberOfPriors": "",
                "priority": "",
                "usePriors": "",
                "createdBy": "",
                "lastModifiedBy": "",
                "displayDescriptorNames": "",
                "displayDicomTags": "",
                "snapshotViews": [],
                "aghDataUri": "",
                "aghDownloadName": _safe_download_filename("", f"hanging_{protocol_id}", ".AGH"),
            })

    def _prio_sort_key(p):
        raw = (p.get("priority") or "").strip()
        try:
            return (0, int(raw), (p.get("name") or "").lower(), p.get("id") or "")
        except Exception:
            return (1, 999999, (p.get("name") or "").lower(), p.get("id") or "")

    hanging_protocols.sort(key=_prio_sort_key)

    name_counter = Counter(
        (r.get("name", "") or "").strip().lower()
        for r in hanging_condition_set_rows
        if (r.get("name", "") or "").strip()
    )
    duplicate_names = sorted([n for n, c in name_counter.items() if c > 1])
    duplicate_name_group_index = {n: idx for idx, n in enumerate(duplicate_names)}
    for r in hanging_condition_set_rows:
        name_key = (r.get("name", "") or "").strip().lower()
        is_dup = bool(name_key and name_counter.get(name_key, 0) > 1)
        r["isDuplicateName"] = is_dup
        r["duplicateNameGroup"] = duplicate_name_group_index.get(name_key) if is_dup else None

    all_conditional_descriptor_ids = sorted({
        (r.get("descriptorId", "") or "").strip()
        for r in hanging_condition_set_rows
        if (r.get("descriptorId", "") or "").strip()
    })
    all_conditional_descriptor_blocks = [
        descriptor_raw_xml.get(did, "")
        for did in all_conditional_descriptor_ids
        if descriptor_raw_xml.get(did, "")
    ]
    hanging_condition_sets_agd_data_uri = _build_agd_data_uri_from_descriptors(all_conditional_descriptor_blocks) if all_conditional_descriptor_blocks else ""
    hanging_condition_sets_agd_download_name = f"conditions-export-{APP_VERSION}.AGD"

    hanging_condition_set_rows.sort(key=lambda x: ((x.get("name", "") or "").lower(), (x.get("condition", "") or "").lower(), (x.get("dicomTagDescription", "") or "").lower(), (x.get("dicomTagNumber", "") or "").lower(), (x.get("value", "") or "").lower()))
    for h in hanging_protocols:
        h["detailPayload"] = {
            "name": h.get("name", ""),
            "active": h.get("active", ""),
            "numberOfPriors": h.get("numberOfPriors", ""),
            "priority": h.get("priority", ""),
            "usePriors": h.get("usePriors", ""),
            "createdBy": h.get("createdBy", ""),
            "lastModifiedBy": h.get("lastModifiedBy", ""),
            "snapshotViews": h.get("snapshotViews", []),
        }

    hanging_inactive_count = sum(
        1 for h in hanging_protocols if (h.get("active", "") or "").strip().lower() != "true"
    )

    standard_prefix = "listtext.lta.worklist.library.standard.bundles#"
    search_prefix = "listtext.lta.worklist.library.search.bundles#"
    scheduled_prefix = "listtext.lta.worklist.library.scheduled.bundles#"
    user_prefix = "listtext.lta.worklist.library.user.bundles#"
    worklist_bundle_map = defaultdict(dict)
    user_worklist_bundle_map = defaultdict(dict)

    for i in all_items:
        name = _normalize_item_name(i.attrib.get("name", ""))
        value = (i.text or "").strip()

        if name.startswith(standard_prefix):
            rest = name[len(standard_prefix):]
            if "." in rest:
                bid, key = rest.split(".", 1)
                worklist_bundle_map[f"standard:{bid}"][key] = value

        if name.startswith(search_prefix):
            rest = name[len(search_prefix):]
            if "." in rest:
                bid, key = rest.split(".", 1)
                worklist_bundle_map[f"search:{bid}"][key] = value

        if name.startswith(scheduled_prefix):
            rest = name[len(scheduled_prefix):]
            if "." in rest:
                bid, key = rest.split(".", 1)
                worklist_bundle_map[f"scheduled:{bid}"][key] = value

        if name.startswith(user_prefix):
            rest = name[len(user_prefix):]
            if "." in rest:
                bid, key = rest.split(".", 1)
                user_worklist_bundle_map[bid][key] = value

    def _extract_query_nodes(query_nodes_xml):
        if not query_nodes_xml:
            return ""
        try:
            qroot = ET.fromstring(query_nodes_xml)
            items = [x.text.strip() for x in qroot.findall("item") if x.text and x.text.strip()]
            return ", ".join(items)
        except Exception:
            return ""

    def _extract_modalities(advanced_filter_xml):
        if not advanced_filter_xml:
            return ""
        candidates = ["CR", "DX", "CT", "MR", "MG", "US", "XA", "NM", "PT", "OT", "RF"]
        found = []
        for m in candidates:
            if re.search(rf"\b{m}\b", advanced_filter_xml):
                found.append(m)
        return ", ".join(found)

    def _extract_timeframe(*xml_texts):
        entries = []
        for xml_text in xml_texts:
            if not xml_text:
                continue
            try:
                xroot = ET.fromstring(xml_text)
            except Exception:
                continue

            for entry in xroot.findall('.//entry'):
                search_type = (entry.attrib.get('search_type', '') or '').strip().lower()
                search_value = (entry.findtext('searchstring', default='') or '').strip()
                if not search_type:
                    continue

                if search_type == 'days_old' and search_value:
                    entries.append(f"letzte {search_value} Tage")
                elif search_type in {'from_to_date', 'between'} and search_value:
                    entries.append(f"Zeitraum {search_value}")
                elif search_type in {'today', 'yesterday'}:
                    entries.append(search_type)

        unique = []
        for e in entries:
            if e not in unique:
                unique.append(e)
        return ", ".join(unique)

    def _summarize_advanced_filter(advanced_filter_xml):
        if not advanced_filter_xml:
            return ""
        try:
            xroot = ET.fromstring(advanced_filter_xml)
        except Exception:
            return advanced_filter_xml[:300]

        grouped = defaultdict(list)
        for entry in xroot.findall('.//entry'):
            search_type = (entry.attrib.get('search_type', '') or '').strip()
            criterion_tag = (entry.attrib.get('criterion_tag', '') or '').strip()
            values = [
                (s.text or '').strip()
                for s in entry.findall('searchstring')
                if (s.text or '').strip()
            ]
            if not search_type and not criterion_tag and not values:
                continue

            tag_label = DICOM_TAG_LABELS.get(criterion_tag, criterion_tag)
            key = (search_type, tag_label)
            grouped[key].extend(values)

        parts = []
        for (search_type, tag_label), vals in grouped.items():
            dedup = []
            for v in vals:
                if v not in dedup:
                    dedup.append(v)
            short_vals = ", ".join(dedup[:8])
            if len(dedup) > 8:
                short_vals += " …"
            head = f"{search_type}({tag_label})"
            parts.append(f"{head}: {short_vals}" if short_vals else head)

        return " · ".join(parts)

    worklist_bundles = []
    scheduled_worklist_bundles = []
    for full_id, vals in sorted(worklist_bundle_map.items(), key=lambda x: x[0]):
        kind, bid = full_id.split(":", 1)
        role_paths = sorted({v for k, v in vals.items() if ".roles#" in k and k.endswith(".path") and v})
        query_nodes_xml = vals.get("signature.queryNodesXML", "")
        advanced_filter_xml = vals.get("signature.advancedFilterXML", "")
        row = {
            "bundleId": bid,
            "title": vals.get("signature.title", ""),
            "type": vals.get("signature.type", "") or kind.upper(),
            "roles": ", ".join(role_paths),
            "queryNodes": _extract_query_nodes(query_nodes_xml),
            "modality": _extract_modalities(advanced_filter_xml),
            "timeframe": _extract_timeframe(
                advanced_filter_xml,
                vals.get("signature.additionalFilterXML", "")
            ),
            "filterName": vals.get("signature.filter_name", ""),
            "searchType": vals.get("signature.searchType", ""),
        }
        if kind == "standard":
            worklist_bundles.append(row)
        elif kind == "scheduled":
            sched_days = sorted({
                v for k, v in vals.items()
                if k.startswith("signature.schedule.scheduledDays#") and k.endswith(".day") and v
            })
            row["ownerName"] = vals.get("signature.ownerName", "")
            row["scheduledDays"] = ", ".join(sched_days)
            row["advancedFilterSummary"] = _summarize_advanced_filter(advanced_filter_xml)
            scheduled_worklist_bundles.append(row)

    user_worklist_bundle_rows = []
    ignored_user_worklist_bundles = 0
    for bid, vals in sorted(user_worklist_bundle_map.items(), key=lambda x: x[0]):
        owner_id = vals.get("signature.ownerId", "")
        owner_name = vals.get("signature.ownerName", "")
        title = vals.get("signature.title", "")
        # In manchen Exports bleiben vereinzelte verwaiste Fragment-Items übrig
        # (z. B. nur columnsXML ohne owner/title). Diese sind keine auswertbaren
        # User-Bundles und werden für die Profil-Tabelle nicht gezählt.
        if not (owner_id or owner_name or title):
            ignored_user_worklist_bundles += 1
            continue

        role_paths = sorted({v for k, v in vals.items() if ".roles#" in k and k.endswith(".path") and v}, key=lambda x: x.lower())
        query_nodes_xml = vals.get("signature.queryNodesXML", "")
        advanced_filter_xml = vals.get("signature.advancedFilterXML", "")
        user_worklist_bundle_rows.append({
            "configBundleId": bid,
            "bundleId": vals.get("bundleId", ""),
            "ownerId": owner_id,
            "ownerName": owner_name,
            "title": title,
            "type": vals.get("signature.type", "") or "USER",
            "roles": ", ".join(role_paths),
            "queryNodes": _extract_query_nodes(query_nodes_xml),
            "modality": _extract_modalities(advanced_filter_xml),
            "timeframe": _extract_timeframe(
                advanced_filter_xml,
                vals.get("signature.additionalFilterXML", "")
            ),
            "filterName": vals.get("signature.filter_name", ""),
            "searchType": vals.get("signature.searchType", ""),
        })

    user_profile_groups = defaultdict(list)
    for row in user_worklist_bundle_rows:
        owner_key = (row.get("ownerId") or row.get("ownerName") or "(ohne Owner)").strip()
        user_profile_groups[owner_key].append(row)

    user_worklist_profiles = []
    for owner_key, rows in user_profile_groups.items():
        rows_sorted = sorted(rows, key=lambda r: ((r.get("title") or "").lower(), str(r.get("bundleId") or ""), str(r.get("configBundleId") or "")))
        owner_ids = [r.get("ownerId", "") for r in rows_sorted]
        owner_names = [r.get("ownerName", "") for r in rows_sorted]
        titles = [r.get("title", "") for r in rows_sorted]
        all_titles = _join_unique(titles, sep=" | ")
        title_excerpt_values = []
        for t in titles:
            if t and t not in title_excerpt_values:
                title_excerpt_values.append(t)
            if len(title_excerpt_values) >= 12:
                break
        title_excerpt = " | ".join(title_excerpt_values)
        if len([t for t in titles if t]) > len(title_excerpt_values):
            title_excerpt += " …" if title_excerpt else "…"

        roles_text = _join_unique(sorted({r.get("roles", "") for r in rows_sorted if r.get("roles", "")}, key=lambda x: x.lower()), sep=" | ")
        query_nodes_text = _join_unique(sorted({r.get("queryNodes", "") for r in rows_sorted if r.get("queryNodes", "")}, key=lambda x: x.lower()), sep=" | ")
        modality_text = _join_unique(sorted({r.get("modality", "") for r in rows_sorted if r.get("modality", "")}, key=lambda x: x.lower()), sep=" | ")
        timeframe_text = _join_unique(sorted({r.get("timeframe", "") for r in rows_sorted if r.get("timeframe", "")}, key=lambda x: x.lower()), sep=" | ")
        search_type_text = _join_unique(sorted({r.get("searchType", "") for r in rows_sorted if r.get("searchType", "")}, key=lambda x: x.lower()), sep=" | ")

        search_text = " ".join([
            owner_key,
            _join_unique(owner_ids, sep=" "),
            _join_unique(owner_names, sep=" "),
            all_titles,
            roles_text,
            query_nodes_text,
            modality_text,
            timeframe_text,
            search_type_text,
        ])

        user_worklist_profiles.append({
            "ownerId": _join_unique(owner_ids, sep=" | ") or owner_key,
            "ownerName": _join_unique(owner_names, sep=" | "),
            "bundleCount": len(rows_sorted),
            "searchTypes": search_type_text,
            "queryNodes": query_nodes_text,
            "modality": modality_text,
            "timeframe": timeframe_text,
            "roles": roles_text,
            "titleExcerpt": title_excerpt,
            "allTitles": all_titles,
            "searchText": search_text,
        })

    user_worklist_profiles.sort(key=lambda r: (-int(r.get("bundleCount") or 0), (r.get("ownerId") or "").lower(), (r.get("ownerName") or "").lower()))

    user_worklist_stats = {
        "raw_bundles_total": len(user_worklist_bundle_map),
        "valid_bundles_total": len(user_worklist_bundle_rows),
        "ignored_bundles_total": ignored_user_worklist_bundles,
        "owners_total": len(user_worklist_profiles),
    }

    workflow_map = defaultdict(dict)
    workflow_roles_map = defaultdict(set)
    workflow_prefix = "pasta.workflows#"
    for i in all_items:
        name = _normalize_item_name(i.attrib.get("name", ""))
        if not name.startswith(workflow_prefix):
            continue
        rest = name[len(workflow_prefix):]
        if "." not in rest:
            continue
        wid, key = rest.split(".", 1)
        workflow_map[wid][key] = (i.text or "").strip()

    for r in roles:
        rp = r.get("path", "")
        for item_node in r.get("item_nodes", []):
            name = _normalize_item_name(item_node.attrib.get("name", ""))
            if not name.startswith(workflow_prefix):
                continue
            rest = name[len(workflow_prefix):]
            if "." not in rest:
                continue
            wid, _ = rest.split(".", 1)
            if rp:
                workflow_roles_map[wid].add(rp)

    import_workflows = []
    export_workflows = []
    import_prop_keys = set()
    export_prop_keys = set()
    for wid, vals in sorted(workflow_map.items(), key=lambda x: x[0]):
        wtype = (vals.get("type", "") or "").strip()
        wtype_low = wtype.lower()
        if "import" not in wtype_low and "export" not in wtype_low and "burn" not in wtype_low:
            continue

        props_raw = vals.get("properties", "") or ""
        props = {}
        for token in props_raw.split():
            if "=" not in token:
                continue
            k, v = token.split("=", 1)
            props[k] = v

        row = {
            "id": wid,
            "name": vals.get("name", ""),
            "type": wtype,
            "roles": ", ".join(sorted(workflow_roles_map.get(wid, set()))),
            "enabled": vals.get("enabled", ""),
            "includeInContextMenu": vals.get("includeInContextMenu", ""),
            "props": props,
        }

        if "import" in wtype_low:
            import_workflows.append(row)
            import_prop_keys.update(props.keys())
        if "export" in wtype_low or "burn" in wtype_low:
            export_workflows.append(row)
            export_prop_keys.update(props.keys())

    import_workflow_columns = sorted(import_prop_keys)
    export_workflow_columns = sorted(export_prop_keys)
    import_workflow_column_labels = {k: WORKFLOW_PROP_LABELS.get(k, k) for k in import_workflow_columns}
    export_workflow_column_labels = {k: WORKFLOW_PROP_LABELS.get(k, k) for k in export_workflow_columns}

    mcs_map = defaultdict(dict)
    mcs_prefix = "listtext.export.dicommediacreation.servers#"
    for i in all_items:
        name = _normalize_item_name(i.attrib.get("name", ""))
        if not name.startswith(mcs_prefix):
            continue
        rest = name[len(mcs_prefix):]
        if "." not in rest:
            continue
        sid, key = rest.split(".", 1)
        mcs_map[sid][key] = (i.text or "").strip()

    media_creation_servers = []
    for sid, vals in sorted(mcs_map.items(), key=lambda x: x[0]):
        media_creation_servers.append({
            "id": sid,
            "name": vals.get("name", ""),
            "host": vals.get("host", ""),
            "port": vals.get("port", ""),
            "calledAET": vals.get("calledAET", ""),
            "use_cmove": vals.get("use_cmove", ""),
            "notSendImages": vals.get("notSendImages", ""),
            "medium": vals.get("medium", ""),
        })

    dicom_printer_map = defaultdict(dict)
    dp_prefix = "listtext.print.dicom_printers#"
    for i in all_items:
        name = _normalize_item_name(i.attrib.get("name", ""))
        if not name.startswith(dp_prefix):
            continue
        rest = name[len(dp_prefix):]
        if "." not in rest:
            continue
        pid, key = rest.split(".", 1)
        dicom_printer_map[pid][key] = (i.text or "").strip()

    def _extract_attr(raw_xml: str, tag: str, attr: str):
        if not raw_xml:
            return ""
        m = re.search(rf"<{tag}[^>]*\b{attr}=\"([^\"]*)\"", raw_xml)
        return m.group(1) if m else ""

    def _extract_media(raw_xml: str):
        if not raw_xml:
            return ""
        media_items = re.findall(r"<item\s+size=\"([^\"]+)\"\s+type=\"([^\"]+)\"", raw_xml)
        return "; ".join([f"{size} {mtype}" for size, mtype in media_items])

    dicom_printers_by_mcs = []
    for pid, vals in sorted(dicom_printer_map.items(), key=lambda x: x[0]):
        printer_xml = (vals.get("xml", "") or "").replace("&lt;", "<").replace("&gt;", ">")
        dicom_printers_by_mcs.append({
            "id": pid,
            "media_creation_server": vals.get("name", ""),
            "host": vals.get("host", ""),
            "port": vals.get("port", ""),
            "calledAET": vals.get("calledAET", ""),
            "callingAET": vals.get("callingAET", ""),
            "destination": _extract_attr(printer_xml, "destination", "value"),
            "resolution": _extract_attr(printer_xml, "resolution", "value"),
            "media": _extract_media(printer_xml),
            "format": _extract_attr(printer_xml, "item", "format"),
            "color": _extract_attr(printer_xml, "sop_classes", "color"),
            "annotation": _extract_attr(printer_xml, "sop_classes", "annotation"),
            "presentation_lut": _extract_attr(printer_xml, "sop_classes", "presentation_lut"),
            "overlay": _extract_attr(printer_xml, "sop_classes", "overlay"),
        })

    def _dup_norm(v: str) -> str:
        return (v or "").strip().lower()

    printer_dup_counter = Counter(
        (
            _dup_norm(p.get("host")),
            _dup_norm(p.get("calledAET")),
        )
        for p in dicom_printers_by_mcs
        if _dup_norm(p.get("host")) and _dup_norm(p.get("calledAET"))
    )

    duplicate_keys = sorted([k for k, c in printer_dup_counter.items() if c > 1])
    duplicate_group_index = {k: idx for idx, k in enumerate(duplicate_keys)}

    for p in dicom_printers_by_mcs:
        dup_key = (
            _dup_norm(p.get("host")),
            _dup_norm(p.get("calledAET")),
        )
        is_dup = bool(dup_key[0] and dup_key[1] and printer_dup_counter.get(dup_key, 0) > 1)
        p["isDuplicate"] = is_dup
        p["duplicateGroup"] = duplicate_group_index.get(dup_key) if is_dup else None

    # Keep printers with the same duplicate color/group next to each other too.
    dicom_printers_by_mcs.sort(key=lambda p: (
        0 if p.get("isDuplicate") else 1,
        p.get("duplicateGroup") if p.get("duplicateGroup") is not None else 999999,
        _dup_norm(p.get("host")),
        _dup_norm(p.get("calledAET")),
        (p.get("media_creation_server") or "").strip().lower(),
        str(p.get("id") or ""),
    ))

    workstation_stats = {
        "roles_total": len(workstation_roles),
        "roles_with_items": sum(1 for r in workstation_roles if r["items"] > 0),
        "items_total": len(workstation_item_details),
        "raw_devices_total": len(raw_workstation_devices),
        "devices_total": len(workstation_devices),
        "duplicate_groups_total": sum(1 for d in workstation_devices if d.get("hasDuplicate")),
        "stale_devices_total": sum(1 for d in workstation_devices if d["isStale"]),
    }

    return {
        "xml_kind": "exportxml",
        "root": root.tag,
        "roles_count": len(roles),
        "items_count": len(all_items),
        "type_counts": sorted(type_counts.items(), key=lambda x: x[0]),
        "enterprise_roles": enterprise_roles,
        "workstation_roles": workstation_roles,
        "workstation_stats": workstation_stats,
        "enterprise_permission_rows": enterprise_permission_rows,
        "enterprise_permission_roles": enterprise_permission_roles,
        "workstation_item_details": workstation_item_details,
        "workstation_devices": workstation_devices,
        "workstation_license_columns": workstation_license_columns,
        "workstation_license_rows": workstation_license_rows,
        "dicom_endpoints": dicom_endpoints,
        "archives": archives,
        "scripts": scripts,
        "hanging_protocols": hanging_protocols,
        "hanging_condition_sets": hanging_condition_set_rows,
        "hanging_condition_sets_agd_data_uri": hanging_condition_sets_agd_data_uri,
        "hanging_condition_sets_agd_download_name": hanging_condition_sets_agd_download_name,
        "hanging_stats": {
            "inactive_count": hanging_inactive_count,
        },
        "worklist_bundles": worklist_bundles,
        "scheduled_worklist_bundles": scheduled_worklist_bundles,
        "user_worklist_profiles": user_worklist_profiles,
        "user_worklist_bundle_rows": user_worklist_bundle_rows,
        "user_worklist_stats": user_worklist_stats,
        "import_workflows": import_workflows,
        "import_workflow_columns": import_workflow_columns,
        "import_workflow_column_labels": import_workflow_column_labels,
        "export_workflows": export_workflows,
        "export_workflow_columns": export_workflow_columns,
        "export_workflow_column_labels": export_workflow_column_labels,
        "media_creation_servers": media_creation_servers,
        "dicom_printers_by_mcs": dicom_printers_by_mcs,
        "csv_hits": csv_hits,
    }


def _require_api_key():
    configured = (os.getenv("API_KEY") or "").strip()
    if not configured:
        return None

    provided = (request.headers.get("X-API-Key") or "").strip()
    if provided != configured:
        return jsonify({"error": "Unauthorized"}), 401
    return None


def _require_script_access_code_api_enabled():
    if not _is_script_access_code_api_enabled():
        abort(404)


def _script_access_status_payload():
    return {
        "feature": "script_access_code",
        "enabled": _is_script_access_code_enabled(),
        "default_enabled_from_env": bool(SCRIPT_ACCESS_CODE_DEFAULT_ENABLED),
        "env_var": "SCRIPT_ACCESS_CODE_ENABLED",
        "api_enabled": _is_script_access_code_api_enabled(),
        "api_default_enabled_from_env": bool(SCRIPT_ACCESS_CODE_API_DEFAULT_ENABLED),
        "api_env_var": "SCRIPT_ACCESS_CODE_API_ENABLED",
        "note": "Runtime changes through this API are reset to SCRIPT_ACCESS_CODE_ENABLED after container restart. Set SCRIPT_ACCESS_CODE_API_ENABLED=false in Docker Compose to disable this API.",
    }


@app.get("/api/v1/script-access-code/status")
def api_script_access_code_status():
    _require_script_access_code_api_enabled()
    auth_err = _require_api_key()
    if auth_err is not None:
        return auth_err
    return jsonify(_script_access_status_payload())


@app.post("/api/v1/script-access-code/status")
def api_set_script_access_code_status():
    _require_script_access_code_api_enabled()
    auth_err = _require_api_key()
    if auth_err is not None:
        return auth_err

    payload = request.get_json(silent=True) or {}
    if "enabled" not in payload:
        return jsonify({"error": "Missing JSON field: enabled"}), 400

    enabled = _parse_bool(payload.get("enabled"), None)
    if enabled is None:
        return jsonify({"error": "enabled must be boolean or one of true/false/on/off/1/0"}), 400

    _set_script_access_code_enabled(enabled)
    return jsonify(_script_access_status_payload())


# Short aliases for convenience.
@app.get("/api/v1/script-access/status")
def api_script_access_status_alias():
    return api_script_access_code_status()


@app.post("/api/v1/script-access/status")
def api_set_script_access_status_alias():
    return api_set_script_access_code_status()


@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "GET":
        return render_template("index.html")

    files = [f for f in request.files.getlist("xml_file") if f and (f.filename or "").strip()]
    if not files:
        return render_template("index.html", error="Bitte eine XML-Datei, ein Lizenzfile oder performance.csv auswählen.")
    if len(files) > 2:
        return render_template("index.html", error="Bitte maximal zwei Dateien auswählen. Der Vergleich ist nur für zwei performance.csv-Dateien vorgesehen.")

    filenames = [(f.filename or "") for f in files]
    lower_filenames = [name.lower() for name in filenames]
    if not all(name.endswith(".xml") or name.endswith(".csv") for name in lower_filenames):
        return render_template("index.html", error="Bitte nur XML- oder CSV-Dateien auswählen.")

    if len(files) == 2:
        if not all(name.endswith(".csv") for name in lower_filenames):
            return render_template("index.html", error="Der Vergleich mit zwei Dateien ist nur für performance.csv-Dateien möglich. export.xml und Lizenzfile bitte einzeln analysieren.")
        try:
            parsed = []
            for f, filename in zip(files, filenames):
                raw = f.read()
                if not raw:
                    return render_template("index.html", error=f"Datei ist leer oder konnte nicht gelesen werden: {filename}")
                parsed.append(analyze_performance_csv(BytesIO(raw), filename=filename))
            result = _build_performance_comparison(parsed[0], parsed[1], filenames[0], filenames[1])
        except Exception:
            return render_template("index.html", error="Fehler beim Parsen der performance.csv-Dateien. Bitte CSV-Struktur prüfen.")
        detected_type = "Performance CSV Vergleich"
        filename = f"{filenames[0]} ↔ {filenames[1]}"
        return render_template("index.html", result=result, filename=filename, detected_type=detected_type)

    f = files[0]
    filename = filenames[0]
    lower_filename = lower_filenames[0]

    raw = f.read()
    if not raw:
        return render_template("index.html", error="Datei ist leer oder konnte nicht gelesen werden.")

    if lower_filename.endswith(".csv"):
        try:
            result = analyze_performance_csv(BytesIO(raw), filename=filename)
        except Exception:
            return render_template("index.html", error="Fehler beim Parsen der performance.csv. Bitte CSV-Struktur prüfen.")
    else:
        try:
            result = analyze_xml(BytesIO(raw))
        except Exception:
            return render_template("index.html", error="Fehler beim Parsen der Datei. Bitte XML-Struktur prüfen.")

    if result.get("xml_kind") == "licensefile":
        detected_type = "Lizenzfile"
    elif result.get("xml_kind") == "performancecsv":
        detected_type = "Performance CSV"
    else:
        detected_type = "Konfigurationsfile"

    return render_template("index.html", result=result, filename=filename, detected_type=detected_type)


@app.get("/hanging/<protocol_id>")
def hanging_detail(protocol_id):
    # Hanging details are intentionally no longer cached server-side.
    # Detail pages are generated client-side from the currently displayed browser result.
    abort(404)


@app.get("/anleitung")
def anleitung():
    return render_template("anleitung.html")


@app.get("/script_access_code")
def script_access_code():
    if not _is_script_access_code_enabled():
        abort(404)
    now = datetime.now()
    code = _generate_script_access_code()
    return render_template(
        "script_access_code.html",
        code=code,
        generated_at=now.strftime("%d.%m.%Y %H:%M:%S"),
    )


@app.get("/script_access_code_json")
def script_access_code_json():
    if not _is_script_access_code_enabled():
        return jsonify({"enabled": False, "error": "ScriptAccess Code is disabled"}), 404
    now = datetime.now()
    now_ts = int(now.timestamp() * 1000)
    c_now = _generate_script_access_code(now_ts)
    m20 = now.replace(minute=20, second=0, microsecond=0)
    c_m20 = _generate_script_access_code(int(m20.timestamp() * 1000))
    m_40 = now.replace(minute=0, second=0, microsecond=0) - timedelta(minutes=40)
    c_m40 = _generate_script_access_code(int(m_40.timestamp() * 1000))
    return jsonify({
        "enabled": True,
        "code": c_now,
        "codes": [c_now, c_m20, c_m40],
        "generated_at": now.strftime("%d.%m.%Y %H:%M:%S"),
    })


@app.post("/api/v1/export-xml/analyze")
def api_analyze_export_xml():
    auth_err = _require_api_key()
    if auth_err is not None:
        return auth_err

    f = request.files.get("file")
    raw = None
    source_filename = None

    if f is not None:
        source_filename = f.filename or "export.xml"
        raw = f.read()
    else:
        raw = request.get_data(cache=False)
        source_filename = "export.xml"

    if not raw:
        return jsonify({"error": "Leere oder fehlende XML-Datei."}), 400

    try:
        result = analyze_xml(BytesIO(raw))
    except Exception:
        return jsonify({"error": "Fehler beim Parsen der Datei. Bitte XML-Struktur prüfen."}), 400

    if result.get("xml_kind") != "exportxml":
        return jsonify({"error": "Es wird eine export.xml (Konfigurationsfile) erwartet."}), 400

    return jsonify({
        "status": "ok",
        "filename": source_filename,
        "xml_kind": "exportxml",
        "summary": {
            "roles_count": result.get("roles_count", 0),
            "items_count": result.get("items_count", 0),
            "archives_count": len(result.get("archives") or []),
            "dicom_endpoints_count": len(result.get("dicom_endpoints") or []),
            "scripts_count": len(result.get("scripts") or []),
            "workstations_count": len(result.get("workstation_devices") or []),
        },
    })


@app.post("/api/v1/performance/analyze")
def api_analyze_performance_csv():
    auth_err = _require_api_key()
    if auth_err is not None:
        return auth_err

    f = request.files.get("file")
    raw = None
    source_filename = None

    if f is not None:
        source_filename = f.filename or "performance.csv"
        raw = f.read()
    else:
        raw = request.get_data(cache=False)
        source_filename = "performance.csv"

    if not raw:
        return jsonify({"error": "Leere oder fehlende performance.csv."}), 400

    try:
        result = analyze_performance_csv(BytesIO(raw), filename=source_filename)
    except Exception:
        return jsonify({"error": "Fehler beim Parsen der performance.csv. Bitte CSV-Struktur prüfen."}), 400

    return jsonify({
        "status": "ok",
        "filename": source_filename,
        "xml_kind": "performancecsv",
        "summary": result.get("performance_summary", {}),
        "modalities": result.get("performance_modalities", []),
        "users": result.get("performance_users", []),
        "hangup_counts": result.get("performance_hangup_counts", []),
    })


@app.get("/download_bundle")
def download_bundle():
    if not _is_deploy_download_enabled():
        abort(404)

    app_dir = Path(__file__).resolve().parent

    compose_content = """services:
  export-xml-web:
    build:
      context: ./webapp
    image: export-xml-web:latest
    container_name: export-xml-web
    ports:
      - "18080:8080"
    environment:
      API_KEY: "CHANGE_ME_TO_A_LONG_RANDOM_SECRET"
      # Set to "false" to hide and disable ScriptAccess Code.
      SCRIPT_ACCESS_CODE_ENABLED: "true"
      # Set to "false" to disable the runtime API for changing ScriptAccess Code status.
      SCRIPT_ACCESS_CODE_API_ENABLED: "true"
      # Set to "false" to hide and disable Deploy Download.
      DEPLOY_DOWNLOAD_ENABLED: "true"
    restart: unless-stopped
"""

    compose_image_content = """services:
  export-xml-web:
    image: ghcr.io/syschelle/export-xml-web:latest
    container_name: export-xml-web
    ports:
      - "18080:8080"
    environment:
      API_KEY: "CHANGE_ME_TO_A_LONG_RANDOM_SECRET"
      # Set to "false" to hide and disable ScriptAccess Code.
      SCRIPT_ACCESS_CODE_ENABLED: "true"
      # Set to "false" to disable the runtime API for changing ScriptAccess Code status.
      SCRIPT_ACCESS_CODE_API_ENABLED: "true"
      # Set to "false" to hide and disable Deploy Download.
      DEPLOY_DOWNLOAD_ENABLED: "true"
    restart: unless-stopped
"""

    github_workflow_content = """name: Build Docker images

"on":
  push:
    branches:
      - main
      - master
    tags:
      - "v*"
  workflow_dispatch:

permissions:
  contents: read
  packages: write

env:
  REGISTRY: ghcr.io
  IMAGE_NAME: export-xml-web
  APP_VERSION: v0.161

jobs:
  build-export-xml-web:
    name: Build export-xml-web image
    runs-on: ubuntu-latest

    steps:
      - name: Checkout repository
        uses: actions/checkout@v4

      - name: Prepare Docker image tags
        id: image
        shell: bash
        run: |
          owner="${GITHUB_REPOSITORY_OWNER,,}"
          image="${REGISTRY}/${owner}/${IMAGE_NAME}"
          short_sha="${GITHUB_SHA::12}"

          {
            echo "image=${image}"
            echo "tags<<EOF"
            echo "${image}:${APP_VERSION}"
            echo "${image}:sha-${short_sha}"
            if [[ "${GITHUB_REF_TYPE}" == "tag" ]]; then
              echo "${image}:${GITHUB_REF_NAME}"
            fi
            echo "${image}:latest"
            echo "EOF"
          } >> "${GITHUB_OUTPUT}"

      - name: Set up Docker Buildx
        uses: docker/setup-buildx-action@v4

      - name: Login to GitHub Container Registry
        uses: docker/login-action@v4
        with:
          registry: ${{ env.REGISTRY }}
          username: ${{ github.actor }}
          password: ${{ secrets.GITHUB_TOKEN }}

      - name: Build and push Docker image
        uses: docker/build-push-action@v7
        with:
          context: ./webapp
          file: ./webapp/Dockerfile
          platforms: linux/amd64
          push: true
          tags: ${{ steps.image.outputs.tags }}
          labels: |
            org.opencontainers.image.title=export-xml-web
            org.opencontainers.image.description=ConfigScope
            org.opencontainers.image.source=${{ github.server_url }}/${{ github.repository }}
            org.opencontainers.image.revision=${{ github.sha }}
            org.opencontainers.image.version=${{ env.APP_VERSION }}
          cache-from: type=gha
          cache-to: type=gha,mode=max

      - name: Show compose image values
        shell: bash
        run: |
          echo "Image gebaut: ${{ steps.image.outputs.image }}:${APP_VERSION}"
          echo "docker-compose.image.yml starten mit:"
          echo "docker compose -f docker-compose.image.yml up -d"
"""

    install_en_content = """# Installation Guide (Docker)

This guide explains how to run ConfigScope on an external system.

## Requirements
- Docker Engine / Docker Desktop
- Docker Compose v2 (`docker compose`)
- For the GitHub workflow: GitHub Actions with package write permissions enabled

## API key (recommended)
Set `API_KEY` directly in `docker-compose.yml` or `docker-compose.image.yml`.
Use the same value in request header `X-API-Key`.
If `API_KEY` is empty or missing, API access is open.

No `.env` file is required.

## ScriptAccess Code feature flags
Set the ScriptAccess options directly in the Compose file.

```yaml
SCRIPT_ACCESS_CODE_ENABLED: "true"       # show and enable ScriptAccess Code
SCRIPT_ACCESS_CODE_ENABLED: "false"      # hide and disable ScriptAccess Code
SCRIPT_ACCESS_CODE_API_ENABLED: "true"   # allow the runtime API toggle
SCRIPT_ACCESS_CODE_API_ENABLED: "false"  # disable the runtime API toggle
```

`SCRIPT_ACCESS_CODE_ENABLED` controls the UI/button and `/script_access_code` endpoints.
`SCRIPT_ACCESS_CODE_API_ENABLED` controls the API endpoints used to read or change the ScriptAccess Code status at runtime.

After changing the Compose file, recreate the container:

```bash
docker compose -f docker-compose.image.yml up -d --force-recreate
```

Runtime API toggle, protected by `X-API-Key` when `API_KEY` is configured and only available when `SCRIPT_ACCESS_CODE_API_ENABLED` is `"true"`:

```bash
curl -H "X-API-Key: <your-key>" http://localhost:18080/api/v1/script-access-code/status
curl -X POST http://localhost:18080/api/v1/script-access-code/status \
  -H "X-API-Key: <your-key>" \
  -H "Content-Type: application/json" \
  -d '{"enabled": false}'
```

Runtime API changes are reset to the Docker Compose `SCRIPT_ACCESS_CODE_ENABLED` value after container restart. If `SCRIPT_ACCESS_CODE_API_ENABLED` is `"false"`, the runtime API endpoints return `404`.

## Deploy Download feature flag
Set the Deploy Download option directly in the Compose file.

```yaml
DEPLOY_DOWNLOAD_ENABLED: "true"   # show and enable Deploy Download
DEPLOY_DOWNLOAD_ENABLED: "false"  # hide and disable Deploy Download
```

When disabled, the Deploy Download button is hidden and `/download_bundle` returns `404`.
After changing the Compose file, recreate the container.

## Local start with build from source
Use this variant when Docker should build the image locally from `webapp/Dockerfile`.

```bash
docker compose up -d --build
```

## Start with prebuilt GitHub Container Registry image
Use this variant when the image was built by GitHub Actions and pushed to GitHub Container Registry.

```bash
docker compose -f docker-compose.image.yml up -d
```

Image used by `docker-compose.image.yml`:

```text
ghcr.io/syschelle/export-xml-web:latest
```

## GitHub Actions image build
The workflow is stored here:

```text
.github/workflows/docker-images.yml
```

It builds the Docker image from:

```text
webapp/Dockerfile
```

The workflow pushes these tags to GitHub Container Registry:

- `v0.161`
- `sha-<short-sha>`
- `latest` for the current published image
- the Git tag name when a `v*` tag is pushed

`docker-compose.image.yml` uses `ghcr.io/syschelle/export-xml-web:latest` by default. Edit the `image:` line if you want to pin a fixed version tag.

## Open
- http://localhost:18080
- or http://<SERVER-IP>:18080

## API example
```bash
curl -X POST "http://localhost:18080/api/v1/export-xml/analyze" \
  -H "X-API-Key: <your-key>" \
  -F "file=@/path/to/export.xml;type=application/xml"
```


## Data handling and demo disclaimer
ConfigScope is intended only as a demonstration and analysis aid. Use is at the user's own risk.

Uploaded XML, license and performance CSV files are processed in memory for the current request. The application does not intentionally write uploaded files, parsed analysis results or generated report data to disk, a database or persistent server-side storage. After the response has been rendered, the displayed analysis exists in the browser only.

Exports such as Excel, PDF/print output, AGH/AGD files and the deploy bundle are generated on demand and are stored only where the user explicitly saves or downloads them. Standard infrastructure such as browser caches, reverse proxies, web server access logs or container/platform logging can still record technical metadata such as request paths, timestamps or client addresses depending on the deployment environment.

## Stop
```bash
docker compose down
# or, for the image based compose file:
docker compose -f docker-compose.image.yml down
```
"""

    buf = BytesIO()
    with zipfile.ZipFile(buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("docker-compose.yml", compose_content)
        zf.writestr("docker-compose.image.yml", compose_image_content)
        zf.writestr(".github/workflows/docker-images.yml", github_workflow_content)
        zf.writestr("INSTALL_EN.md", install_en_content)

        for p in app_dir.rglob("*"):
            if not p.is_file():
                continue
            rel = p.relative_to(app_dir)
            rel_str = str(rel)
            if "__pycache__" in rel.parts:
                continue
            if rel_str.endswith(".pyc"):
                continue
            if Path(rel_str).name.lower().startswith("merged_export") and rel_str.lower().endswith(".xml"):
                continue
            zf.write(p, arcname=f"webapp/{rel_str}")

    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/zip",
        as_attachment=True,
        download_name=f"export-xml-deploy-bundle-{APP_VERSION}.zip",
    )


@app.post("/export_excel")
def export_excel():
    payload = request.get_json(silent=True) or {}
    sheets = payload.get("sheets") or []

    if not isinstance(sheets, list) or not sheets:
        return jsonify({"error": "Keine Tabellen ausgewählt."}), 400

    wb = Workbook()
    wb.remove(wb.active)

    used_names = set()

    def safe_sheet_name(name):
        cleaned = re.sub(r"[\\/*?:\[\]]", "_", (name or "").strip()) or "Tabelle"
        cleaned = cleaned[:31]
        base = cleaned
        n = 2
        while cleaned in used_names:
            suffix = f"_{n}"
            cleaned = (base[:31 - len(suffix)] + suffix)[:31]
            n += 1
        used_names.add(cleaned)
        return cleaned

    for idx, sheet in enumerate(sheets, start=1):
        title = safe_sheet_name(sheet.get("title") or f"Tabelle {idx}")
        headers = sheet.get("headers") or []
        rows = sheet.get("rows") or []

        ws = wb.create_sheet(title=title)
        if headers:
            ws.append([str(x) for x in headers])
            for cell in ws[1]:
                cell.font = Font(bold=True)

        for row in rows:
            ws.append(["" if v is None else str(v) for v in row])

        for col in ws.columns:
            max_len = 0
            for cell in col:
                val = "" if cell.value is None else str(cell.value)
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 60)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = (payload.get("filename") or "export-xml.xlsx").strip() or "export-xml.xlsx"
    if not filename.lower().endswith(".xlsx"):
        filename += ".xlsx"

    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@app.post("/export_scripts_bundle")
def export_scripts_bundle():
    payload = request.get_json(silent=True) or {}
    scripts = payload.get("scripts") or []
    excel_payload = payload.get("excel") or {}

    if not isinstance(scripts, list) or not scripts:
        return jsonify({"error": "Keine Scripts vorhanden."}), 400

    sheets = excel_payload.get("sheets") or []
    if not isinstance(sheets, list) or not sheets:
        return jsonify({"error": "Excel-Daten fehlen."}), 400

    # Excel via bestehender Logik erzeugen
    wb = Workbook()
    wb.remove(wb.active)
    used_names = set()

    def safe_sheet_name(name):
        cleaned = re.sub(r"[\\/*?:\[\]]", "_", (name or "").strip()) or "Tabelle"
        cleaned = cleaned[:31]
        base = cleaned
        n = 2
        while cleaned in used_names:
            suffix = f"_{n}"
            cleaned = (base[:31 - len(suffix)] + suffix)[:31]
            n += 1
        used_names.add(cleaned)
        return cleaned

    for idx, sheet in enumerate(sheets, start=1):
        title = safe_sheet_name(sheet.get("title") or f"Tabelle {idx}")
        headers = sheet.get("headers") or []
        rows = sheet.get("rows") or []

        ws = wb.create_sheet(title=title)
        if headers:
            ws.append([str(x) for x in headers])
            for cell in ws[1]:
                cell.font = Font(bold=True)

        for row in rows:
            ws.append(["" if v is None else str(v) for v in row])

        for col in ws.columns:
            max_len = 0
            for cell in col:
                val = "" if cell.value is None else str(cell.value)
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 60)

    excel_buf = BytesIO()
    wb.save(excel_buf)
    excel_buf.seek(0)

    zip_buf = BytesIO()
    with zipfile.ZipFile(zip_buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("scripts_export.xlsx", excel_buf.getvalue())

        for idx, s in enumerate(scripts, start=1):
            script_name = (s.get("name") or f"script_{idx}").strip()
            safe_base = re.sub(r"[^A-Za-z0-9._-]+", "_", script_name).strip("._-") or f"script_{idx}"

            code = s.get("fullCode")
            if code:
                zf.writestr(f"scripts/{safe_base}.bsh", code)

            svg_text = s.get("svgText")
            if svg_text:
                zf.writestr(f"svg/{safe_base}.svg", svg_text)

    zip_buf.seek(0)
    return send_file(
        zip_buf,
        mimetype="application/zip",
        as_attachment=True,
        download_name=f"scripts-svg-excel-{APP_VERSION}.zip",
    )


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8080)
